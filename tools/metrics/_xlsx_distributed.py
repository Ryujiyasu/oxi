# -*- coding: utf-8 -*-
"""Where does Excel put the characters of a distributed cell?

The renderer spreads them from edge to edge — the first glyph against the left
gutter, the last against the right. Excel puts a single character in the
middle instead (`_xlsx_indent.py`), which edge-to-edge cannot do, so the rule
needs measuring against two, three and more.

    python tools\\metrics\\_xlsx_distributed.py
"""
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_distributed")
BOOK = SCRATCH / "distributed.xlsx"

TEXTS = ["あ", "あい", "あいう", "あいうえ", "AB", "ABCDE",
         "（①－②）", "(1-2)", "①②", "A①B", "第３表", "あA", "10%", "A B",
         "あ、い", "アイウ", "ＡＢ"]
WIDTHS = [12.0, 20.0]


def build(indent=0):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    for index, width in enumerate(WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    cases, row = [], 1
    for text in TEXTS:
        for column, _ in enumerate(WIDTHS, start=1):
            cell = sheet.cell(row=row, column=column, value=text)
            cell.font = Font(name="ＭＳ Ｐゴシック", size=11)
            cell.alignment = Alignment(horizontal="distributed", vertical="center",
                                       indent=indent)
        sheet.row_dimensions[row].height = 18.0
        cases.append((row, text))
        row += 1
    book.save(BOOK)
    return cases


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def geometry():
    import os

    ours = SCRATCH / "distributed.oxi.png"
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(ours), "96"],
                          capture_output=True, timeout=300, env=environment)
    heights, columns = {}, {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
    return ours, heights, columns


def glyphs(image, top, foot, left, right):
    band = (image[top:foot, left:right] < 128).sum(axis=0)
    out, start = [], None
    for index, lit in enumerate(band > 0):
        if lit and start is None:
            start = index
        elif not lit and start is not None:
            out.append((start, index - 1))
            start = None
    if start is not None:
        out.append((start, len(band) - 1))
    return out


def main():
    indent = int(sys.argv[1]) if len(sys.argv) > 1 else 0
    print(f"indent {indent}")
    cases = build(indent)
    picture = shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, columns = geometry()
    ours = np.asarray(Image.open(ours_png).convert("L"))

    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = (at, at + heights[index])
        at += heights[index]
    across, at = [0], 0
    for index in sorted(columns):
        at += columns[index]
        across.append(at)

    for row, text in cases:
        top, foot = edges[row]
        for column in range(len(WIDTHS)):
            left, right = across[column], across[column + 1]
            theirs = glyphs(truth, top, foot, left, right)
            mine = glyphs(ours, top, foot, left, right)
            print(f"{text:<10}{f'{right - left}px':>7}  Excel {str(theirs):<58} ours {mine}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
