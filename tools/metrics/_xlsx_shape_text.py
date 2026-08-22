# -*- coding: utf-8 -*-
"""How far apart are the lines of a shape's text?

The renderer lays a shape's text on the line box a *cell* would use — the
measured `row_defaults` table — and `002`'s side panel says that is wrong:
メイリオ 14pt comes out 36.6px a line in Excel where the cell box is 30. This
puts a rectangle of three lines on a sheet for each face and size, in a
workbook written by hand because openpyxl has no shapes, and reads the pitch
off Excel's own picture.

    python tools\\metrics\\_xlsx_shape_text.py
    python tools\\metrics\\_xlsx_shape_text.py --reuse
"""
import argparse
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
SCRATCH = Path(r"C:\tmp\xlsx_shape_text")
BOOK = SCRATCH / "shape_text.xlsx"

# Each shape gets a column of the sheet and three lines of text.
CASES = [("メイリオ", 9.0), ("メイリオ", 11.0), ("メイリオ", 12.0), ("メイリオ", 14.0),
         ("メイリオ", 16.0), ("メイリオ", 20.0),
         ("Yu Gothic UI", 12.0), ("游ゴシック", 11.0), ("游ゴシック", 12.0), ("游ゴシック", 14.0),
         ("游ゴシック", 18.0), ("游ゴシック", 36.0),
         ("AR P丸ゴシック体E", 12.0), ("AR P丸ゴシック体E", 14.0), ("AR P丸ゴシック体E", 16.0),
         ("ＤＦ特太ゴシック体", 14.0), ("ＭＳ 明朝", 7.0), ("ＭＳ 明朝", 11.0),
         ("ＭＳ Ｐゴシック", 11.0), ("ＭＳ Ｐゴシック", 14.0), ("Meiryo UI", 11.0),
         ("Calibri", 11.0), ("Calibri", 18.0), ("Arial", 11.0),
         # The same faces again with Latin text — 10.5 and 13.5 point stand
         # in for "write Hxpq here", which is how the builder tells them apart.
         ("メイリオ", 13.5), ("ＭＳ Ｐゴシック", 13.5), ("游ゴシック", 13.5)]
# The same letters on every line: a line's ink top depends on which
# glyphs are on it, so the pitch can only be read off lines that hold
# the same ones.
LINES = ["国国国国", "国国国国", "国国国国"]
# Rows a shape is given. Three lines of 36pt need room, and a shape's
# text must never reach the next one's band.
SPACING = 16


def anchors():
    """One shape a row, each hung from its own cell."""
    held = []
    for index, (face, points) in enumerate(CASES):
        # A CJK string in a Latin face is drawn by whatever Windows falls
        # back to, and would measure that font instead of this one.
        latin = face in ("Calibri", "Arial", "Times New Roman") or points in (10.5, 13.5)
        runs = "".join(
            f'<a:p><a:pPr algn="l"/><a:r>'
            f'<a:rPr lang="ja-JP" sz="{int(points * 100)}">'
            f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/></a:rPr>'
            f"<a:t>{'Hxpq' if latin else line}</a:t></a:r></a:p>"
            for line in LINES
        )
        held.append(
            f"<xdr:twoCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING + SPACING - 1}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="none" anchor="t"/><a:lstStyle/>{runs}</xdr:txBody>'
            f"</xdr:sp><xdr:clientData/></xdr:twoCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 20.0
    sheet.column_dimensions["C"].width = 20.0
    # Something in the far corner, so Excel's used range covers every shape.
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(CASES) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors()}</xdr:wsDr>"
    )
    sheet_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )

    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            body = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                body = body.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                body = body.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>'
                )
                if "xmlns:r=" not in body:
                    body = body.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ',
                        1,
                    )
                body = body.encode("utf-8")
            out.writestr(item, body)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", sheet_rels)
    shutil.copy(BOOK, SCRATCH / "shape_text_copy.xlsx")


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    height, width = truth.shape
    print(f"picture {width}x{height}")

    # Each shape's three lines are the only ink in its band of rows.
    ink = (truth < 128).sum(axis=1)
    bands, start = [], None
    for y, lit in enumerate(ink > 0):
        if lit and start is None:
            start = y
        elif not lit and start is not None:
            bands.append((start, y - 1))
            start = None
    lines = [band for band in bands if band[1] - band[0] >= 3]
    # Each shape owns a stretch of the picture: the rows it was given.
    row_px = height / (len(CASES) * SPACING + 2)

    print(f"{'face':<16}{'pt':>6}{'em px':>7}{'line tops':>22}{'pitch':>8}{'/em':>7}")
    for index, (face, points) in enumerate(CASES):
        floor, ceiling = index * SPACING * row_px, (index + 1) * SPACING * row_px
        three = [band for band in lines if floor <= band[0] < ceiling][:3]
        if len(three) < 3:
            print(f"{face:<16}{points:>6}   (only {len(three)} lines found)")
            continue
        tops = [band[0] for band in three]
        pitch = (tops[2] - tops[0]) / 2
        em = points * 96 / 72
        print(f"{face:<16}{points:>6}{em:>7.1f}{str(tops):>22}{pitch:>8.1f}{pitch / em:>7.3f}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
