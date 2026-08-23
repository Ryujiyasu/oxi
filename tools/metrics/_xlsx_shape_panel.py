# -*- coding: utf-8 -*-
r"""What is it about `sanko_tool`'s panel that lets its line stay whole?

`_xlsx_shape_room.py` sweeps a plain box under that panel's own line and finds
Excel's turn at a room of 557.75 — but the panel's room is 557.0 and Excel
keeps the line whole there anyway. So something the panel carries and the
swept box does not is worth a pixel of room.

Each arm here is the same line in a box of exactly the panel's width, with one
more of the panel's own properties than the arm before it: its rounded
geometry, its three-point rule, its middle anchor, its clipping, and finally
its whole paragraph with the breaks inside it.

    python tools\metrics\_xlsx_shape_panel.py
    python tools\metrics\_xlsx_shape_panel.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_panel")
BOOK = SCRATCH / "panel.xlsx"
EMU = 9525.0

SPACING = 12
FACE, POINTS = "游ゴシック", 12.0
# The panel's own box, to the pixel Excel states it in: 432.15 point wide.
WIDTH = round(432.15 * 4 / 3 * EMU)
LINE = ("   「調査票番号」、「品目番号」、「アイテム記号」をプルダウンで選択する")
HEAD = "１．確認したい品目アイテムについて、「品目アイテム指定」シートから"

ARMS = [
    ("plain", dict()),
    ("round", dict(geometry="roundRect")),
    ("ruled", dict(geometry="roundRect", rule=True)),
    ("middle", dict(geometry="roundRect", rule=True, anchor="ctr")),
    ("clipped", dict(geometry="roundRect", rule=True, anchor="ctr", clip=True)),
    ("whole", dict(geometry="roundRect", rule=True, anchor="ctr", clip=True,
                   paragraph=True)),
    # And the panel's own words in a plain box, to tell the paragraph apart
    # from everything else it carries.
    ("paragraph only", dict(paragraph=True)),
]


def body(arm):
    size = int(POINTS * 100)
    dress = (f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>')
    if arm.get("paragraph"):
        runs = (f'<a:p><a:pPr algn="l"/>'
                f'<a:r><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr>'
                f"<a:t>イメージ</a:t></a:r>"
                f'<a:br><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr></a:br>'
                f'<a:br><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr></a:br>'
                f'<a:r><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr>'
                f"<a:t>{HEAD}</a:t></a:r>"
                f'<a:br><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr></a:br>'
                f'<a:r><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr>'
                f"<a:t>{LINE}</a:t></a:r></a:p>")
    else:
        runs = (f'<a:p><a:pPr algn="l"/>'
                f'<a:r><a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr>'
                f"<a:t>{LINE}</a:t></a:r></a:p>")
    over = ' vertOverflow="clip" horzOverflow="clip"' if arm.get("clip") else ""
    anchor = arm.get("anchor", "t")
    return (f'<xdr:txBody><a:bodyPr{over} wrap="square" rtlCol="0" '
            f'anchor="{anchor}"/><a:lstStyle/>{runs}</xdr:txBody>')


def anchors_xml():
    held = []
    for index, (_name, arm) in enumerate(ARMS):
        geometry = arm.get("geometry", "rect")
        rule = ('<a:ln w="38100"><a:solidFill><a:srgbClr val="0000FF"/></a:solidFill></a:ln>'
                if arm.get("rule") else "<a:ln><a:noFill/></a:ln>")
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{WIDTH}" cy="{int(11 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="{geometry}"><a:avLst/></a:prstGeom>'
            f'<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>{rule}</xdr:spPr>'
            f"{body(arm)}</xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 90.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(ARMS) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "panel.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def how_many(picture, top, foot, lane):
    """How many lines of words the band holds, ignoring the box's own rule."""
    # Well inside the box: its own sides light every row, and a rounded
    # corner lights the ones near the top and foot.
    held = (picture[top:foot, lane + 20:lane + 540] < 128).sum(axis=1)
    lines, run = 0, 0
    for lit in held:
        # A rule across the box lights the whole width; words light a little.
        if lit > 0:
            run += 1
        else:
            if run > 3:
                lines += 1
            run = 0
    return lines + (1 if run > 3 else 0)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"the box is {WIDTH / EMU:.2f} pixels wide, the line 557.75 of run")
    print(f"{'arm':<16}{'Excel lines':>12}{'ours':>6}")
    for index, (name, _arm) in enumerate(ARMS):
        top = edges.get(index * SPACING + 1)
        foot = edges.get(index * SPACING + SPACING)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = how_many(truth, top, foot, lane)
        ours = how_many(mine, top, foot, lane)
        print(f"{name:<16}{theirs:>12}{ours:>6}{'' if theirs == ours else '  <<'}")


if __name__ == "__main__":
    main()
