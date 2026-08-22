# -*- coding: utf-8 -*-
"""Does Excel keep a row's stated height when the row is not pinned?

The renderer recomputes every row that is not `customHeight`, which is right
for 25723 of the corpus's 25736 rows. `h2dee1989kre` row 5 is one of the
thirteen it is wrong about: the row states `ht="20.25"` — 27px — with no pin,
its cells are 游ゴシック 11 whose own line is 25px, and Excel draws 27. That
one row costs the workbook 0.27 of SSIM, because everything below it is two
pixels out.

This asks Excel what it does with a stated height that is not pinned, against
a height above the row's own, below it, and equal to it, with the row empty
and with text in it.

    python tools\\metrics\\_xlsx_stated_height.py
"""
import re
import shutil
import sys
import zipfile
from pathlib import Path

import win32com.client

SCRATCH = Path(r"C:\tmp\xlsx_stated_height")
BOOK = SCRATCH / "stated.xlsx"

# (stated height in points, whether the row is pinned, what the row holds)
# The last field names the dressing: plain text, text in a ruled cell, text
# centred in its cell, and so on — `h2dee1989kre` row 5 is ruled and centred
# and comes out taller than the same font on its own.
CASES = [
    (20.25, False, "text"),
    (20.25, False, "empty"),
    (20.25, True, "text"),
    (15.0, False, "text"),
    (30.0, False, "text"),
    (None, False, "text"),
    (None, False, "ruled"),
    (None, False, "centred"),
    (None, False, "ruled and centred"),
    (None, False, "filled"),
    (None, False, "latin"),
    (None, False, "ruled empty"),
]


def build(normal=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    # The workbook this is chasing wears 游ゴシック 11 as its Normal font, and
    # a row of that font in it comes out 27px where the same font in a
    # Calibri workbook comes out 25.
    if normal:
        face, points = normal
        book._named_styles["Normal"].font = Font(name=face, size=points)
    sheet = book.active
    from openpyxl.styles import Alignment, Border, PatternFill, Side
    edge = Side(style="thin", color="FF000000")
    frame = Border(left=edge, right=edge, top=edge, bottom=edge)
    for index, (height, _pinned, holds) in enumerate(CASES, start=1):
        if holds != "empty":
            text = "abc" if holds == "latin" else "あいう"
            cell = sheet.cell(row=index, column=1,
                              value=None if holds == "ruled empty" else text)
            cell.font = Font(name="游ゴシック", size=11)
            if "ruled" in holds:
                cell.border = frame
            if "centred" in holds:
                cell.alignment = Alignment(vertical="center")
            if holds == "filled":
                cell.fill = PatternFill("solid", fgColor="FFFFCC")
        if height is not None:
            sheet.row_dimensions[index].height = height
    book.save(plain)

    # openpyxl pins every height it writes; the question is what an unpinned
    # one does, so the flag is taken back out where the case says so.
    with zipfile.ZipFile(plain) as source, zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            body = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = body.decode("utf-8")
                for index, (_height, pinned, _holds) in enumerate(CASES, start=1):
                    if pinned:
                        continue
                    text = re.sub(
                        rf'(<row r="{index}"[^>]*?) customHeight="1"',
                        r"\1",
                        text,
                    )
                body = text.encode("utf-8")
            out.writestr(item, body)
    shutil.copy(BOOK, SCRATCH / "stated_copy.xlsx")


def main():
    normal = None
    if len(sys.argv) > 1:
        face, points = sys.argv[1].rsplit(":", 1)
        normal = (face, float(points))
        print(f"Normal font: {face} {points}")
    build(normal)
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        workbook = excel.Workbooks.Open(str(BOOK.resolve()), ReadOnly=True)
        sheet = workbook.Worksheets(1)
        print(f"{'stated':>8}{'pinned':>8}{'holds':>20}{'Excel px':>10}")
        for index, (height, pinned, holds) in enumerate(CASES, start=1):
            drawn = sheet.Rows(index).Height * 96 / 72
            print(f"{str(height):>8}{str(pinned):>8}{holds:>20}{drawn:>10.1f}")
        workbook.Close(SaveChanges=False)
    finally:
        excel.Quit()


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
