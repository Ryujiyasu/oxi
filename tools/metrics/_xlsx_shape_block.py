# -*- coding: utf-8 -*-
r"""How tall is a shape's block of text, and where in the box does it sit?

`sanko_tool`'s イメージ panel is anchored `ctr` and sets its whole block about
a line higher than Excel does, which is what still holds shape text behind its
flag. A block one line too tall hangs its first line above the box — so the
question is how Excel counts lines: a paragraph, a `<a:br/>` inside one, an
empty paragraph between two, an empty one at the end.

Each case is a shape of a known height holding a known structure, drawn with
no fill and no rule so the only ink is the text. What is read back is where
the ink starts and ends inside the box, and how many lines it holds.

    python tools\metrics\_xlsx_shape_block.py
    python tools\metrics\_xlsx_shape_block.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_block")
BOOK = SCRATCH / "block.xlsx"

POINTS = 12.0
SPACING = 8          # rows per case, so each shape has a band of its own
FACE = "ＭＳ Ｐゴシック"
WORD = "あA"

# Each spec is a list of paragraphs; a paragraph is a list of lines, and a
# paragraph of no lines is the empty one Excel writes between blocks.
SPECS = [
    ("4 paragraphs", [[WORD], [WORD], [WORD], [WORD]]),
    ("1 para, 3 br", [[WORD, WORD, WORD, WORD]]),
    ("2 + empty + 1", [[WORD], [WORD], [], [WORD]]),
    ("empty first", [[], [WORD], [WORD], [WORD]]),
    ("empty last", [[WORD], [WORD], [WORD], []]),
    ("two empties", [[WORD], [], [], [WORD]]),
    ("br + empty para", [[WORD, WORD], [], [WORD, WORD]]),
    ("one line", [[WORD]]),
]
ANCHORS = ["t", "ctr", "b"]


def body(spec, anchor):
    held = []
    for lines in spec:
        runs = ""
        if not lines:
            runs = (f'<a:endParaRPr lang="ja-JP" sz="{int(POINTS * 100)}">'
                    f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>'
                    f"</a:endParaRPr>")
        else:
            for index, line in enumerate(lines):
                if index:
                    runs += ("<a:br><a:rPr lang=\"ja-JP\" "
                             f'sz="{int(POINTS * 100)}"><a:latin typeface="{FACE}"/>'
                             f'<a:ea typeface="{FACE}"/></a:rPr></a:br>')
                runs += (f'<a:r><a:rPr lang="ja-JP" sz="{int(POINTS * 100)}">'
                         f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>'
                         f"</a:rPr><a:t>{line}</a:t></a:r>")
        held.append(f'<a:p><a:pPr algn="l"/>{runs}</a:p>')
    return (f'<xdr:txBody><a:bodyPr wrap="square" anchor="{anchor}"/>'
            f'<a:lstStyle/>{"".join(held)}</xdr:txBody>')


def cases():
    held = []
    for name, spec in SPECS:
        for anchor in ANCHORS:
            held.append((name, spec, anchor))
    return held


def anchors_xml():
    held = []
    for index, (name, spec, anchor) in enumerate(cases()):
        held.append(
            f"<xdr:twoCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING + SPACING - 1}</xdr:row>"
            f"<xdr:rowOff>0</xdr:rowOff></xdr:to>"
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f"{body(spec, anchor)}</xdr:sp><xdr:clientData/></xdr:twoCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 20.0
    sheet.column_dimensions["C"].width = 20.0
    # A corner in column A so the used range starts at the top left; the scan
    # skips that column, so it cannot be read as a shape's ink.
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=900)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "block.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=900,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def bands_of(picture, top, foot, lane):
    """Each run of lit scanlines in the band, as (first, last)."""
    held = (picture[top:foot, lane:] < 128).sum(axis=1)
    lines, run = [], None
    for step, lit in enumerate(held):
        if lit:
            run = [step, step] if run is None else [run[0], step]
        elif run is not None and step - run[1] > 2:
            lines.append(tuple(run))
            run = None
    if run is not None:
        lines.append(tuple(run))
    return lines


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"{'structure':<16}{'anchor':>7}{'box':>6}"
          f"{'Excel lines':>12}{'top':>6}{'foot':>6}"
          f"{'ours lines':>12}{'top':>6}{'foot':>6}   ")
    for index, (name, _spec, anchor) in enumerate(cases()):
        # The shape hangs from its own row and stops a row short of the next.
        top = edges.get(index * SPACING)
        foot = edges.get(index * SPACING + SPACING - 1)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = bands_of(truth, top, foot, lane)
        ours = bands_of(mine, top, foot, lane)
        flag = "" if (len(theirs) == len(ours)
                      and theirs[:1] == ours[:1]) else "  <<"
        print(f"{name:<16}{anchor:>7}{foot - top:>6}"
              f"{len(theirs):>12}{theirs[0][0] if theirs else -1:>6}"
              f"{theirs[-1][1] if theirs else -1:>6}"
              f"{len(ours):>12}{ours[0][0] if ours else -1:>6}"
              f"{ours[-1][1] if ours else -1:>6}{flag}")


if __name__ == "__main__":
    main()
