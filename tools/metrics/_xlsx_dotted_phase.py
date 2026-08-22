# -*- coding: utf-8 -*-
"""Where does the phase of a dotted rule come from, column by column?

A dotted rule across the sheet is inked where `(x + 2y) mod 4` is 0 or 3 —
that fits every horizontal rule measured. The rules running down the sheet
do not follow it: of the four in the first probe, three read one phase and
one the other, and neither `x mod 4` nor `x mod 8` separates them. This puts
twenty rules down the sheet at deliberately scattered x and asks which of
them are in phase, so the law can be read off the list.

    python tools\\metrics\\_xlsx_dotted_phase.py
    python tools\\metrics\\_xlsx_dotted_phase.py --reuse
"""
import argparse
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_border")
BOOK = SCRATCH / "dotted_phase.xlsx"

# Twenty columns whose widths walk the boundary through every residue.
WIDTHS = [2.0, 2.14, 2.29, 2.43, 2.57, 2.71, 2.86, 3.0, 3.14, 3.29,
          3.43, 3.57, 3.71, 3.86, 4.0, 4.14, 4.29, 4.43, 4.57, 4.71]


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    side = Side(style="dotted", color="FF000000")
    for column, width in enumerate(WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in range(1, 4):
        sheet.row_dimensions[row].height = 18.0
        for column in range(1, len(WIDTHS) + 1):
            sheet.cell(row=row, column=column).border = Border(left=side, right=side)
    book.save(BOOK)
    return BOOK


def shoot(path):
    picture = path.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{path.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=300)
    listing.unlink(missing_ok=True)
    return picture


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    path = build()
    picture = path.with_suffix(".excel.png") if args.reuse else shoot(path)
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    height, width = truth.shape
    print(f"picture {width}x{height}")
    print(f"{'x':>5}{'width':>7}{'dark y mod 4':>16}{'(x+y) mod 4':>14}{'x mod 4':>9}{'x mod 8':>9}")
    for x in range(width):
        dark = [y for y in range(height) if truth[y, x] < 128]
        if len(dark) < height // 3:          # not a rule: a crossing or nothing
            continue
        residues = sorted({y % 4 for y in dark})
        print(f"{x:>5}{'':>7}{str(residues):>16}"
              f"{str(sorted({(x + y) % 4 for y in dark})):>14}{x % 4:>9}{x % 8:>9}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
