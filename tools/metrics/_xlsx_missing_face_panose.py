# -*- coding: utf-8 -*-
"""Which installed face Excel puts in place of one this machine has not.

`sanko_tool` asks its callouts for `AR P丸ゴシック体E`, which is not installed.
Asked for by name alone (`_xlsx_shape_text.py`) Excel draws it at 1.31 line
heights a line — the ＭＳ family's number. In `sanko_tool` itself the same
face at the same size comes out at 27px a line, which is 游ゴシック's. The
difference between the two files is what the run carries beside the name:
`panose`, `pitchFamily` and `charset`. This puts both spellings on a sheet,
with the plausible substitutes beside them, and reads the pitch and the width
of a known line off Excel's own picture.

    python tools\\metrics\\_xlsx_missing_face_panose.py
    python tools\\metrics\\_xlsx_missing_face_panose.py --reuse
"""
import argparse
import os
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_missing_face")
BOOK = SCRATCH / "missing_face.xlsx"

MISSING = "AR P丸ゴシック体E"
# What `sanko_tool`'s runs carry beside the name.
DRESSED = ' panose="020F0900000000000000" pitchFamily="50" charset="-128"'
CASES = [
    ("name only", MISSING, ""),
    ("with panose", MISSING, DRESSED),
    ("ＭＳ Ｐゴシック", "ＭＳ Ｐゴシック", ""),
    ("ＭＳ ゴシック", "ＭＳ ゴシック", ""),
    ("游ゴシック", "游ゴシック", ""),
    ("メイリオ", "メイリオ", ""),
    ("Yu Gothic UI", "Yu Gothic UI", ""),
]
# Kanji are full width in every face; only the kana and the marks
# between them tell a proportional face from a fixed one.
LINES = ["国「あ、い」。国ぁ", "国「あ、い」。国ぁ", "国「あ、い」。国ぁ"]
POINTS = 12.0
SPACING = 8


def anchors():
    held = []
    for index, (_name, face, extra) in enumerate(CASES):
        runs = "".join(
            f'<a:p><a:pPr algn="l"/><a:r>'
            f'<a:rPr lang="ja-JP" sz="{int(POINTS * 100)}">'
            f'<a:latin typeface="{face}"{extra}/>'
            f'<a:ea typeface="{face}"{extra}/></a:rPr>'
            f"<a:t>{line}</a:t></a:r></a:p>"
            for line in LINES
        )
        held.append(
            f"<xdr:twoCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING + SPACING - 1}</xdr:row>"
            f"<xdr:rowOff>0</xdr:rowOff></xdr:to>"
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="none" anchor="t"/><a:lstStyle/>{runs}'
            f"</xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 24.0
    sheet.column_dimensions["C"].width = 24.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(CASES) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            body = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                body = body.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                body = body.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>'
                )
                if "xmlns:r=" not in body:
                    body = body.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ',
                        1,
                    )
                body = body.encode("utf-8")
            out.writestr(item, body)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))

    ours = SCRATCH / "missing_face.oxi.png"
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_SHAPE_TEXT="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(ours), "96"],
                          capture_output=True, timeout=300, env=environment)
    heights = {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    mine_png = np.asarray(Image.open(ours).convert("L")) if ours.exists() else None
    print(f"{'asked for':<18}{'band':>12}{'line tops':>22}{'pitch':>7}{'/em':>7}"
          f"{'ink':>6}{'ours':>6}")
    em = POINTS * 96 / 72
    for index, (name, _face, _extra) in enumerate(CASES):
        # Each shape hangs from a row of its own, so its band is known.
        top = edges.get(index * SPACING + 1, 0)
        foot = edges.get((index + 1) * SPACING + 1, truth.shape[0])
        rows = []
        for y in range(top, min(foot, truth.shape[0])):
            if (truth[y] < 128).sum() > 4:
                if rows and y - rows[-1][-1] <= 2:
                    rows[-1].append(y)
                else:
                    rows.append([y])
        if len(rows) < 2:
            print(f"{name:<18}{f'{top}..{foot}':>12}  (not found)")
            continue
        tops = [r[0] for r in rows]
        pitch = (tops[-1] - tops[0]) / (len(tops) - 1)
        xs = np.flatnonzero((truth[rows[0][0]:rows[0][-1] + 1] < 128).sum(axis=0))
        width = int(xs[-1] - xs[0] + 1) if xs.size else 0
        ours_width = 0
        if mine_png is not None and rows[0][-1] < mine_png.shape[0]:
            mx = np.flatnonzero(
                (mine_png[rows[0][0]:rows[0][-1] + 1] < 128).sum(axis=0))
            ours_width = int(mx[-1] - mx[0] + 1) if mx.size else 0
        print(f"{name:<18}{f'{top}..{foot}':>12}{str(tops):>22}{pitch:>7.1f}"
              f"{pitch / em:>7.3f}{width:>6}{ours_width:>6}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
