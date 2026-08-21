# -*- coding: utf-8 -*-
"""Where Excel puts the baseline in a row, at the top, the middle and the bottom.

One character with no descender, so the ink says exactly where the baseline
was: draw the same character with the device at a baseline this program
chooses, and the distance from that baseline to the top of its ink is a
constant that can be subtracted from ink measured in Excel's own picture.

Three columns — top, centre and bottom of the cell — and several row heights
per font, so the box the text sits in can be told apart from the row itself.

    python tools\\metrics\\_xlsx_baseline_probe.py
    python tools\\metrics\\_xlsx_baseline_probe.py --reuse
"""
import argparse
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _xlsx_font_metrics import ink_ascent, measured_rows, metrics  # noqa: E402

REPO = Path(__file__).resolve().parents[2]
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
SCRATCH = Path(r"C:\tmp\xlsx_valign")
BOOK = SCRATCH / "baseline_probe.xlsx"
TRUTH = SCRATCH / "baseline_probe.excel.png"
OURS = SCRATCH / "baseline_probe.oxi.png"

FONTS = [("ＭＳ ゴシック", 9.0), ("ＭＳ ゴシック", 10.0), ("ＭＳ ゴシック", 11.0),
         ("ＭＳ ゴシック", 14.0), ("ＭＳ 明朝", 11.0), ("ＭＳ 明朝", 16.0),
         ("ＭＳ Ｐゴシック", 9.0), ("ＭＳ Ｐ明朝", 12.0), ("ＭＳ ＵＩゴシック", 11.0),
         ("Meiryo UI", 9.0), ("Meiryo UI", 11.0), ("Meiryo UI", 14.0),
         ("メイリオ", 11.0), ("Yu Gothic", 11.0), ("Yu Gothic UI", 11.0),
         ("游ゴシック", 11.0), ("游ゴシック", 16.0), ("游明朝", 11.0),
         ("Calibri", 11.0), ("Calibri", 18.0), ("Arial", 11.0),
         ("Times New Roman", 12.0), ("Segoe UI", 11.0), ("Consolas", 11.0)]
EXTRA = [0, 6, 30]          # pixels added to the font's own row height
PLACES = ["top", "center", "bottom"]
GLYPH = "亜"                 # sits on the baseline, nothing below it
LATIN = "H"


def build(rows_table):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    for column in range(1, len(PLACES) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 8
    probes, row = [], 1
    for face, points in FONTS:
        natural = rows_table.get((face, points))
        for extra in EXTRA:
            letter = LATIN if face in ("Calibri", "Arial") else GLYPH
            for column, place in enumerate(PLACES, start=1):
                cell = sheet.cell(row=row, column=column, value=letter)
                cell.font = Font(name=face, size=points)
                cell.alignment = Alignment(vertical=place, horizontal="left")
            if natural is not None and extra:
                sheet.row_dimensions[row].height = (natural + extra) * 0.75
            probes.append((row, face, points, natural, extra, letter))
            row += 1
    book.save(BOOK)
    return probes


def heights_from_excel(probes):
    """What Excel makes each row, which is what its own picture is banded by."""
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        book = excel.Workbooks.Open(str(BOOK), 0, True)
        sheet = book.Worksheets(1)
        held = {probe[0]: int((float(sheet.Rows(probe[0]).Height) + 0.05) / 0.75)
                for probe in probes}
        book.Close(False)
    finally:
        excel.Quit()
    return held


def shoot():
    listing = SCRATCH / "_batch3.txt"
    listing.write_text(f"{BOOK.resolve()}\t{TRUTH.resolve()}", encoding="utf-8")
    TRUTH.unlink(missing_ok=True)
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=300)
    listing.unlink(missing_ok=True)


def draw():
    """Draws the sheet, returning the row and column geometry Oxi used."""
    import os

    OURS.unlink(missing_ok=True)
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(OURS), "96"],
                          capture_output=True, text=True, encoding="utf-8",
                          errors="replace", timeout=300, env=environment)
    rows, columns = {}, {}
    for line in done.stdout.splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            rows[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
    return rows, columns


def ink_top(image, top, bottom, left, right):
    band = image[top:bottom, left:right]
    dark = np.flatnonzero((band < 128).any(axis=1))
    return None if dark.size == 0 else int(dark[0])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    # What each font makes a row of its own is asked of Excel rather than
    # assumed, so a face the measured table has never seen is still probed at
    # heights of its own plus a known slack.
    first = build({})
    unpinned = heights_from_excel(first)
    naturals = {}
    for row, face, points, _, _, _ in first:
        naturals.setdefault((face, points), unpinned[row])
    probes = build(naturals)
    excel_rows = heights_from_excel(probes)
    if not args.reuse:
        shoot()
    our_rows, our_columns = draw()
    truth = np.asarray(Image.open(TRUTH).convert("L"))
    ours = np.asarray(Image.open(OURS).convert("L"))

    edges, at = [0], 0
    for index in sorted(our_columns):
        at += our_columns[index]
        edges.append(at)

    print(f"{'face':<16}{'pt':>5}{'nat':>5}{'row':>5}{'asc':>4}{'desc':>5}"
          f"{'h':>4}{'int':>4}{'  top:E   O':>13}{'  mid:E   O':>13}"
          f"{'  bot:E   O':>13}{'nat-d-B':>8}")
    excel_at, our_at = 0, 0
    for row, face, points, natural, extra, letter in probes:
        _, tm = metrics(face, points)
        above = ink_ascent(face, points, letter)
        height = excel_rows[row]
        our_height = our_rows.get(row, height)
        own = naturals.get((face, points), height)
        line = (f"{face:<16}{points:>5.1f}{own:>5}{height:>5}"
                f"{tm.tmAscent:>4}{tm.tmDescent:>5}{tm.tmHeight:>4}"
                f"{tm.tmInternalLeading:>4}")
        seen = []
        for column in range(len(PLACES)):
            left, right = edges[column], edges[column + 1]
            excel = ink_top(truth, excel_at, excel_at + height, left, right)
            mine = ink_top(ours, our_at, our_at + our_height, left, right)
            excel = None if excel is None or above is None else excel + above
            mine = None if mine is None or above is None else mine + above
            seen.append(excel)
            line += f"{(excel if excel is not None else '-'):>7}"
            line += f"{(mine if mine is not None else '-'):>6}"
        # The gap Excel leaves under the baseline in a row of the font's own
        # height — the number the renderer needs and the device does not give.
        under = "-" if seen[0] is None else str(own - tm.tmDescent - seen[0])
        print(line + f"{under:>8}")
        excel_at += height
        our_at += our_height


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
