# -*- coding: utf-8 -*-
"""How Excel draws a cell whose text is stacked — `textRotation="255"`.

771 of the 774 rotations in the corpus are this one, in 39 of its 285
workbooks, and the renderer draws them flat, which is what puts data_A28 at
the bottom of the gate. This asks Excel three things:

  * how tall it makes a row of stacked text of N characters,
  * where each character sits, down and across,
  * and what it does when the row is too short to hold them all.

    python tools\\metrics\\_xlsx_stacked_probe.py
    python tools\\metrics\\_xlsx_stacked_probe.py --reuse
"""
import argparse
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
SCRATCH = Path(r"C:\tmp\xlsx_stacked")
BOOK = SCRATCH / "stacked.xlsx"
TRUTH = SCRATCH / "stacked.excel.png"

FONTS = [("ＭＳ Ｐゴシック", 11.0), ("ＭＳ ゴシック", 9.0), ("ＭＳ 明朝", 11.0)]
LENGTHS = [1, 2, 3, 5, 8]
# None leaves the row to Excel; a number pins it, to see what is done with
# text that will not fit.
PINNED = [None, 30, 60]
WIDTH = 6.0             # characters — narrower than one line of the text
SAMPLE = "政府統計コード情報"


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = WIDTH
    sheet.column_dimensions["B"].width = WIDTH
    probes, row = [], 1
    for face, points in FONTS:
        for length in LENGTHS:
            for pinned in PINNED:
                text = SAMPLE[:length]
                # A: stacked. B: the same text left alone, for comparison.
                stacked = sheet.cell(row=row, column=1, value=text)
                stacked.font = Font(name=face, size=points)
                stacked.alignment = Alignment(textRotation=255, vertical="bottom",
                                              horizontal="center")
                plain = sheet.cell(row=row, column=2, value=text)
                plain.font = Font(name=face, size=points)
                if pinned:
                    sheet.row_dimensions[row].height = pinned * 0.75
                probes.append((row, face, points, length, pinned))
                row += 1
    book.save(BOOK)
    return probes


def heights(probes):
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        book = excel.Workbooks.Open(str(BOOK), 0, True)
        sheet = book.Worksheets(1)
        held = {probe[0]: int((float(sheet.Rows(probe[0]).Height) + 0.05) / 0.75)
                for probe in probes}
        book.Close(False)
    finally:
        excel.Quit()
    return held


def shoot():
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{TRUTH.resolve()}", encoding="utf-8")
    TRUTH.unlink(missing_ok=True)
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=300)
    listing.unlink(missing_ok=True)


def rows_of_ink(band):
    """Which runs of rows in this band hold ink — one run per character."""
    dark = (band < 128).any(axis=1)
    runs, start = [], None
    for index, lit in enumerate(dark):
        if lit and start is None:
            start = index
        elif not lit and start is not None:
            runs.append((start, index - 1))
            start = None
    if start is not None:
        runs.append((start, len(dark) - 1))
    return runs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    probes = build()
    tall = heights(probes)
    if not args.reuse:
        shoot()
    truth = np.asarray(Image.open(TRUTH).convert("L"))
    print(f"picture {truth.shape[1]}x{truth.shape[0]}")

    # The stacked column is the first; find its right edge by the gap between
    # the two columns' ink.
    print(f"{'face':<14}{'pt':>5}{'chars':>6}{'pinned':>7}{'row':>5}"
          f"{'ink runs':>10}{'pitch':>7}{'ink x':>12}  runs")
    top = 0
    for row, face, points, length, pinned in probes:
        height = tall[row]
        band = truth[top:top + height, :]
        # Column A only: everything left of the widest gap in the first 90px.
        left_band = band[:, :int(WIDTH * 8 + 10)]
        runs = rows_of_ink(left_band)
        pitch = (runs[1][0] - runs[0][0]) if len(runs) > 1 else 0
        columns = np.flatnonzero((left_band < 128).any(axis=0))
        span = f"{columns[0]}..{columns[-1]}" if columns.size else "-"
        print(f"{face:<14}{points:>5.1f}{length:>6}{str(pinned):>7}{height:>5}"
              f"{len(runs):>10}{pitch:>7}{span:>12}  "
              f"{[f'{a}-{b}' for a, b in runs[:6]]}")
        top += height


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
