# -*- coding: utf-8 -*-
"""Does an indent start from the same place in a merged block?

`h2daa2023_dendeba_kmc` sets its lines in a merged pair with `indent="2"`, and
the renderer puts them six pixels right of Excel's — even though the same font
and the same indent land exactly right in an unmerged cell
(`_xlsx_indent.py`). The difference has to be the merge, so this asks the same
question of both: one column of plain cells, one of merged pairs, indent 0 to
3, and where the ink starts in each.

    python tools\\metrics\\_xlsx_indent_merged.py
    python tools\\metrics\\_xlsx_indent_merged.py --reuse
"""
import argparse
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_indent_merged")
BOOK = SCRATCH / "indent_merged.xlsx"

FONTS = [("游ゴシック", 11.0), ("ＭＳ Ｐゴシック", 11.0), ("ＭＳ Ｐゴシック", 14.0)]
INDENTS = [0, 1, 2, 3]
# A merged block whose text starts with a space is the shape `dendeba` has.
TEXTS = ["あ", " あ", " あいうえおかきくけこさしすせそたちつてとなにぬねの"]
# A plain cell in column A, a merged pair across C and D.
PLAIN, MERGED = 1, 3


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    for column, width in (("A", 24.0), ("B", 2.0), ("C", 12.0), ("D", 12.0)):
        sheet.column_dimensions[column].width = width
    cases, row = [], 1
    for face, points in FONTS:
        for indent in INDENTS:
            for text in TEXTS:
                for column, merged in ((PLAIN, False), (MERGED, True)):
                    cell = sheet.cell(row=row, column=column, value=text)
                    cell.font = Font(name=face, size=points)
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center", indent=indent
                    )
                    if merged:
                        sheet.merge_cells(start_row=row, start_column=MERGED,
                                          end_row=row, end_column=MERGED + 1)
                    sheet.row_dimensions[row].height = 18.0
                    cases.append((row, face, points, indent, merged, text))
                row += 1
    book.save(BOOK)
    return cases


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def geometry():
    ours = SCRATCH / "indent_merged.oxi.png"
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(ours), "96"],
                          capture_output=True, timeout=300, env=environment)
    heights, widths = {}, {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column":
            widths[int(parts[1])] = int(float(parts[3]))
    return ours, heights, widths


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    cases = build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    ours_png, heights, widths = geometry()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours = np.asarray(Image.open(ours_png).convert("L"))

    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = (at, at + heights[index])
        at += heights[index]
    left = {}
    at = 0
    for index in sorted(widths):
        left[index] = at
        at += widths[index]

    print("column lefts:", left)
    print(f"{'font':<16}{'pt':>5}{'indent':>7}{'where':>9}{'text':>10}"
          f"{'Excel ink':>11}{'ours':>10}")
    for row, face, points, indent, merged, text in cases:
        if row not in edges:
            continue
        top, foot = edges[row]
        if foot > min(truth.shape[0], ours.shape[0]):
            continue
        column = (MERGED if merged else PLAIN) - 1
        start = left[column]
        stop = start + widths[column] * (2 if merged else 1)
        band_truth = truth[top:foot, start:stop]
        band_ours = ours[top:foot, start:stop]
        theirs = np.flatnonzero((band_truth < 128).sum(axis=0))
        mine = np.flatnonzero((band_ours < 128).sum(axis=0))
        if theirs.size == 0:
            continue
        print(f"{face:<16}{points:>5}{indent:>7}"
              f"{'merged' if merged else 'plain':>9}{repr(text[:3]):>10}"
              f"{int(theirs[0]):>11}{(int(mine[0]) if mine.size else -1):>10}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
