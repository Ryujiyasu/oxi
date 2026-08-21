# -*- coding: utf-8 -*-
"""Where does Excel break a wrapped line, character by character?

The row heights say the line *count* is right almost always (0.9994 against
Excel COM), but drawing a sheet shows every break, and a break one character
out is visible. This asks Excel directly: a cell of known text in a column of
known width, wrapped, drawn by Excel, and each line's ink measured back to the
character it ends on — by drawing every candidate prefix with the device and
taking the one whose ink matches.

    python tools\\metrics\\_xlsx_wrap_probe.py
    python tools\\metrics\\_xlsx_wrap_probe.py --reuse
"""
import argparse
import ctypes
import subprocess
import sys
from ctypes import wintypes
from pathlib import Path

import numpy as np
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _xlsx_font_metrics import measured_rows, metrics  # noqa: E402

GDI = ctypes.windll.gdi32
USER = ctypes.windll.user32
REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_wrap")
BOOK = SCRATCH / "wrap.xlsx"
TRUTH = SCRATCH / "wrap.excel.png"

FONTS = [("ＭＳ Ｐゴシック", 11.0), ("ＭＳ ゴシック", 11.0), ("游ゴシック", 11.0)]
WIDTHS = [10.0, 14.0, 20.0]          # in the characters Excel counts columns in
SAMPLES = [
    "行政手続におけるオンラインによる本人確認の手法に関するガイドライン",
    "情報システムを、利用者（国民・事業者等）が、その利用に際して",
    "・ID／パスワード等により認証を行うことで、なりすましを防ぐ。",
    "「行政手続」を、オンライン（インターネット等）で行う場合には、",
    "The quick brown fox jumps over the lazy dog and keeps running",
    "https://www.example.go.jp/digital/data/nq6ept000000kbna-att/x.pdf",
]


def run_width(face, points, text):
    """How wide Excel measures this text for wrapping: each character's own
    advance, added up — the same sum the renderer breaks lines by."""
    if not text:
        return 0
    pixels = round(points * 96.0 / 72.0)
    dc = USER.GetDC(None)
    font = GDI.CreateFontW(-pixels, 0, 0, 0, 400, 0, 0, 0, 1, 0, 0, 5, 0, face)
    old = GDI.SelectObject(dc, font)

    class SIZE(ctypes.Structure):
        _fields_ = [("cx", wintypes.LONG), ("cy", wintypes.LONG)]

    total = 0
    for letter in text:
        size = SIZE()
        if GDI.GetTextExtentPoint32W(dc, letter, len(letter), ctypes.byref(size)):
            total += size.cx
    GDI.SelectObject(dc, old)
    GDI.DeleteObject(font)
    USER.ReleaseDC(None, dc)
    return total


def ink_extent(face, points, text):
    """How wide the ink of this text is when the device draws it."""
    if not text:
        return 0
    pixels = round(points * 96.0 / 72.0)
    width, height, base = 4000, 80, 60
    screen = USER.GetDC(None)
    dc = GDI.CreateCompatibleDC(screen)
    bitmap = GDI.CreateCompatibleBitmap(screen, width, height)
    GDI.SelectObject(dc, bitmap)
    GDI.PatBlt(dc, 0, 0, width, height, 0x00F00062)
    font = GDI.CreateFontW(-pixels, 0, 0, 0, 400, 0, 0, 0, 1, 0, 0, 5, 0, face)
    old = GDI.SelectObject(dc, font)
    GDI.SetBkMode(dc, 1)
    GDI.SetTextAlign(dc, 24)
    GDI.TextOutW(dc, 10, base, text, len(text))

    class HEADER(ctypes.Structure):
        _fields_ = [("biSize", wintypes.DWORD), ("biWidth", wintypes.LONG),
                    ("biHeight", wintypes.LONG), ("biPlanes", wintypes.WORD),
                    ("biBitCount", wintypes.WORD), ("biCompression", wintypes.DWORD),
                    ("biSizeImage", wintypes.DWORD), ("biXPelsPerMeter", wintypes.LONG),
                    ("biYPelsPerMeter", wintypes.LONG), ("biClrUsed", wintypes.DWORD),
                    ("biClrImportant", wintypes.DWORD)]

    header = HEADER()
    header.biSize = ctypes.sizeof(HEADER)
    header.biWidth, header.biHeight = width, -height
    header.biPlanes, header.biBitCount = 1, 32
    buffer = (ctypes.c_ubyte * (width * height * 4))()
    GDI.GetDIBits(dc, bitmap, 0, height, buffer, ctypes.byref(header), 0)
    GDI.SelectObject(dc, old)
    GDI.DeleteObject(font)
    GDI.DeleteObject(bitmap)
    GDI.DeleteDC(dc)
    USER.ReleaseDC(None, screen)

    grid = np.frombuffer(buffer, dtype=np.uint8).reshape(height, width, 4)[:, :, 0]
    lit = np.flatnonzero((grid < 128).any(axis=0))
    return 0 if lit.size == 0 else int(lit[-1] - 10 + 1)


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    for column, width in enumerate(WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    probes, row = [], 1
    for face, points in FONTS:
        for text in SAMPLES:
            for column, _ in enumerate(WIDTHS, start=1):
                cell = sheet.cell(row=row, column=column, value=text)
                cell.font = Font(name=face, size=points)
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal="left")
            # Room for as many lines as the narrowest column can need.
            sheet.row_dimensions[row].height = 20 * 20 * 0.75
            probes.append((row, face, points, text))
            row += 1
    book.save(BOOK)
    return probes


def shoot():
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{TRUTH.resolve()}", encoding="utf-8")
    TRUTH.unlink(missing_ok=True)
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=300)
    listing.unlink(missing_ok=True)


def ours(sheet_path):
    """The columns Oxi laid out, and the lines it drew in each cell."""
    import os
    import re

    environment = dict(os.environ, OXI_XLSX_DUMP_COLUMNS="1", OXI_XLSX_DUMP_LINES="1")
    done = subprocess.run([str(RENDERER), str(sheet_path), str(SCRATCH / "wrap.oxi.png"), "96"],
                          capture_output=True, timeout=300, env=environment)
    # The renderer writes UTF-8 whatever the console's code page says, so its
    # output is decoded here rather than by the pipe.
    out = done.stdout.decode("utf-8", "replace")
    err = done.stderr.decode("utf-8", "replace")
    widths, drawn = {}, {}
    for line in out.splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "column":
            widths[int(parts[1])] = int(float(parts[3]))
    for line in err.splitlines():
        found = re.match(r'drawn row (\d+) col (\d+) lines \[(.*)\]$', line)
        if found:
            held = re.findall(r'"((?:[^"\\]|\\.)*)"', found.group(3))
            # Rust's debug format escapes only the quote, the backslash and
            # the control characters; decoding as unicode_escape would mangle
            # every Japanese character in the line.
            def plain(piece):
                return (piece.replace('\\"', '"')
                             .replace("\\n", chr(10))
                             .replace("\\\\", "\\"))

            drawn[(int(found.group(1)), int(found.group(2)))] = [
                plain(piece) for piece in held
            ]
    return widths, drawn


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    probes = build()
    if not args.reuse:
        shoot()
    widths, drawn = ours(BOOK)
    edges, at = [0], 0
    for index in sorted(widths):
        at += widths[index]
        edges.append(at)
    truth = np.asarray(Image.open(TRUTH).convert("L"))
    rows_table = measured_rows()

    print(f"{'face':<14}{'pt':>5}{'col px':>7}  Excel's lines")
    band_top = 0
    for row, face, points, text in probes:
        line_px = rows_table.get((face, points), 20)
        band = 20 * 20                       # the pinned height in pixels
        for column in range(len(WIDTHS)):
            left, right = edges[column], edges[column + 1]
            found = []
            for index in range(20):
                top = band_top + index * line_px
                strip = truth[top:top + line_px, left:right]
                lit = np.flatnonzero((strip < 128).any(axis=0))
                if lit.size == 0:
                    break
                # How far the line's ink reaches from where its pen started:
                # three pixels in from the cell's left edge, the same place
                # `ink_extent` starts its own drawing from.
                found.append(int(lit[-1]) - 3 + 1)
            # Turn each line's ink into the character it ends on. A line that
            # follows a break at a space starts after the space: Excel does not
            # draw it, so a candidate that carries it measures too wide.
            at_char, breaks = 0, []
            for reach in found:
                while at_char < len(text) and text[at_char] == " ":
                    at_char += 1
                best, best_gap = at_char, 10_000
                for stop in range(at_char + 1, len(text) + 1):
                    gap = abs(ink_extent(face, points, text[at_char:stop].rstrip()) - reach)
                    if gap < best_gap:
                        best, best_gap = stop, gap
                    elif gap > best_gap + 30:
                        break
                breaks.append(best)
                at_char = best
                if at_char >= len(text):
                    break
            excel_lines = [text[a:b] for a, b in zip([0] + breaks, breaks)]
            mine = drawn.get((row, column), [])
            # A space at a break draws nothing wherever it is counted, so the
            # two are compared by what they put on the page.
            def plainly(lines):
                return [line.strip() for line in lines]

            agree = ("" if plainly(mine) == plainly(excel_lines)
                     else "   <<< Oxi: " + " | ".join(mine))
            print(f"{face:<14}{points:>5.1f}{right - left:>7}  "
                  f"{' | '.join(excel_lines)}{agree}")
            # What the column's usable width must be for those breaks: every
            # line fitted, and every line plus its next character did not.
            # A break forced by a kinsoku rule fitted with room to spare, so
            # the bounds it gives are not evidence — the tightest lower bound
            # and the loosest upper one together say where the width lies.
            low, high = 0, 10_000
            for at, stop in zip([0] + breaks, breaks):
                if stop >= len(text):
                    break
                held = text[at:stop].rstrip()
                low = max(low, run_width(face, points, held))
                high = min(high, run_width(face, points, text[at:stop + 1].rstrip()))
            if low and high < 10_000:
                print(f"{'':<26}{'':>7}  fits {low} but not {high}: "
                      f"usable is {low}..{high - 1} of {right - left}, "
                      f"so the allowance is {right - left - (high - 1)}"
                      f"..{right - left - low}")
        band_top += band


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
