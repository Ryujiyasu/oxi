# -*- coding: utf-8 -*-
r"""How much room does a shape's text really get, to the fraction of a pixel?

`sanko_tool`'s panel holds a line 558 pixels wide in a box the renderer makes
577 wide with 19.2 of inset — 557.8 of room — so the renderer breaks the last
character onto a line of its own and Excel does not. The margin is under a
pixel, and the box's own edges are fractional: its anchors carry 1280160 and
624840 EMU, which are 134.4 and 65.6 pixels into their columns.

So this asks where the break really falls. Each case is a shape whose width is
set in EMU a quarter of a pixel at a time around the point where a line of
known width stops fitting, and whose left edge is offset by a fraction of a
pixel as well. The width at which Excel first takes two lines is the room.

    python tools\metrics\_xlsx_shape_room.py
    python tools\metrics\_xlsx_shape_room.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_room")
BOOK = SCRATCH / "room.xlsx"
EMU = 9525.0

SPACING = 5
FACE, POINTS = "游ゴシック", 12.0
LETTERS = 10                    # ten full-width characters: 160 pixels of run
RUN = 160.0
INSET = 2 * 91440 / EMU         # the default insets, 19.2 pixels
# A quarter of a pixel at a time, from a box that cannot hold the line to one
# that comfortably can.
STEPS = [round((RUN + INSET + quarter / 4.0) * EMU)
         for quarter in range(-8, 9)]
# And the same widths again with the left edge a fraction of a pixel in, to
# see whether where the box starts changes what fits in it.
OFFSETS = [0, round(0.4 * EMU), round(0.6 * EMU)]


def cases():
    return [(width, offset) for offset in OFFSETS for width in STEPS]


def anchors_xml():
    held = []
    for index, (width, offset) in enumerate(cases()):
        runs = (f'<a:p><a:pPr algn="l"/><a:r>'
                f'<a:rPr lang="ja-JP" sz="{int(POINTS * 100)}">'
                f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>'
                f'</a:rPr><a:t>{"あ" * LETTERS}</a:t></a:r></a:p>')
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>{offset}</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{width}" cy="{int(4 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f"{runs}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "room.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def how_many(picture, top, foot, lane):
    held = (picture[top:foot, lane:] < 128).sum(axis=1)
    lines, run = 0, False
    for lit in held:
        if lit and not run:
            lines, run = lines + 1, True
        elif not lit:
            run = False
    return lines


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"the line is {RUN:.0f} pixels of run, the insets {INSET:.1f}")
    print(f"{'box px':>9}{'left in':>9}{'room':>9}{'Excel lines':>12}{'ours':>6}")
    for index, (width, offset) in enumerate(cases()):
        top = edges.get(index * SPACING + 1)
        foot = edges.get(index * SPACING + SPACING)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = how_many(truth, top, foot, lane)
        ours = how_many(mine, top, foot, lane)
        mark = "" if theirs == ours else "  <<"
        print(f"{width / EMU:>9.2f}{offset / EMU:>9.2f}{width / EMU - INSET:>9.2f}"
              f"{theirs:>12}{ours:>6}{mark}")


if __name__ == "__main__":
    main()
