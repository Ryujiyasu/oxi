# -*- coding: utf-8 -*-
r"""How far does a pinned pitch lift a shape's line off three quarters?

The renderer puts a pinned line's baseline at `line top + 0.75 x pitch - lift`
with `lift = floor(descent - em/4)`, derived on four arms whose fraction was
0.0 or 0.25 — where floor, round and ceil are the same number. Widening
`_xlsx_shape_pitch.py` to arms with a fraction of 0.667 breaks it in both
directions at once:

    メイリオ 10pt  ours right      游ゴシック 10pt  ours a pixel low
    メイリオ 16pt  a pixel low     游ゴシック 16pt  TWO pixels low
    Yu Gothic UI 16pt  ours right

游ゴシック and Yu Gothic UI at 16 point have the same em and the same device
descent (6), and want lifts of 2 and 0. So the rule is not `descent - em/4`
under any rounding, and the sweep it was derived on was too narrow to say so.

This asks the question on its own: one pinned pitch, four faces, eight sizes,
and the lift read straight off Excel's picture against the arithmetic that
predicts it. Two lines an arm is enough — the pitch is not in question here,
only where in it the first baseline sits.

    python tools\metrics\_xlsx_shape_lift.py
    python tools\metrics\_xlsx_shape_lift.py --reuse
"""
import argparse
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_lift")
BOOK = SCRATCH / "lift.xlsx"
EMU = 9525.0

SPACING = 10            # rows a case
LINES = 2
# A letter that sits ON the baseline, so its ink foot IS the baseline and the
# face's ascent never has to be known.
WORD = "A"
FACES = ["メイリオ", "游ゴシック", "Yu Gothic UI", "ＭＳ Ｐゴシック"]
SIZES = [9.0, 10.0, 11.0, 12.0, 14.0, 16.0, 18.0, 20.0]
# One pinned pitch, stated in hundredths of a point. The lift is what is being
# measured and it does not move with the pitch (`_xlsx_shape_pitch.py`: dead
# constant over eight of them), so one is enough and a big one keeps the two
# lines apart.
PITCH = 2100


def cases():
    return [(face, size) for face in FACES for size in SIZES]


def anchors_xml():
    held = []
    for index, (face, points) in enumerate(cases()):
        one = (
            f'<a:p><a:pPr algn="l">'
            f'<a:lnSpc><a:spcPts val="{PITCH}"/></a:lnSpc></a:pPr>'
            f'<a:r><a:rPr lang="ja-JP" sz="{int(points * 100)}">'
            f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
            f"</a:rPr><a:t>{WORD}</a:t></a:r></a:p>"
        )
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(200 * EMU)}" cy="{int(9 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f"{one * LINES}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "lift.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_DUMP_BLOCK="1", OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    # Where the renderer put each block, so the reading can be turned into a
    # lift rather than a row number.
    # Only the BLOCK's own line: the per-line debug prints an `at=` as well,
    # and matching both interleaves them and puts every block out of step.
    tops = [float(found.group(1))
            for found in re.finditer(r"^block area .*? at=([0-9.]+)",
                                     done.stderr.decode("utf-8", "replace"),
                                     re.M)]
    return ours, heights, lane, tops


def feet(picture, top, foot, lane):
    """The last row of ink of each line — for `A`, the baseline."""
    held = (picture[top:foot, lane:] < 128).sum(axis=1)
    found, run = [], None
    for step, lit in enumerate(held):
        if lit:
            run = step
        elif run is not None and step - run > 1:
            found.append(top + run)
            run = None
    if run is not None:
        found.append(top + run)
    return found


def metrics(face, pixels):
    import win32con
    import win32gui
    import win32ui

    screen = win32gui.GetDC(0)
    dc = win32ui.CreateDCFromHandle(screen)
    font = win32ui.CreateFont({"name": face, "height": -pixels,
                               "weight": win32con.FW_NORMAL,
                               "charset": win32con.DEFAULT_CHARSET})
    old = dc.SelectObject(font)
    told = dc.GetTextMetrics()
    dc.SelectObject(old)
    win32gui.ReleaseDC(0, screen)
    return told


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane, tops = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    pitch_px = PITCH / 100.0 * 96.0 / 72.0
    print(f"pinned pitch {PITCH / 100:.0f}pt = {pitch_px:.2f}px, "
          f"0.75 of it = {0.75 * pitch_px:.2f}")
    print(f"{'face':<14}{'pt':>6}{'em':>8}{'desc':>6}{'exact desc':>12}"
          f"{'asc':>6}{'height':>8}{'Excel lift':>12}{'ours lift':>11}")
    for index, (face, points) in enumerate(cases()):
        top = edges.get(index * SPACING)
        foot = edges.get(index * SPACING + SPACING - 1)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = feet(truth, top, foot, lane)
        ours = feet(mine, top, foot, lane)
        if not theirs or not ours or index >= len(tops):
            continue
        em = points * 96.0 / 72.0
        told = metrics(face, round(em))
        big = metrics(face, 2048)
        # lift = 0.75 x pitch - (baseline - block top), read off each picture.
        block = tops[index]
        print(f"{face:<14}{points:>6}{em:>8.3f}{told['tmDescent']:>6}"
              f"{big['tmDescent'] / 2048 * em:>12.3f}{told['tmAscent']:>6}"
              f"{told['tmHeight']:>8}"
              f"{0.75 * pitch_px - (theirs[0] - block):>12.2f}"
              f"{0.75 * pitch_px - (ours[0] - block):>11.2f}")


if __name__ == "__main__":
    main()
