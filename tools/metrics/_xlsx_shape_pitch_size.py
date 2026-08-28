# -*- coding: utf-8 -*-
r"""What pitch does Excel step a shape's lines by, measured over EIGHT of them?

`_xlsx_shape_pitch.py` sets four lines and reads each step, which resolves a
pitch to a third of a pixel — and the pitch this renderer is wrong by on
`tb_r8_jizensoudan` is a quarter of one. Four lines cannot see it: the panel
there shows it only because 14 + 2 x 26.772 lands on the far side of a
rounding boundary while 14 + 26.772 does not.

So: eight lines of the SAME letter, one face and size a case, and the pitch
read as (last top - first top) / 7, which resolves to a seventh of a pixel.
Ours is read from `OXI_XLSX_DUMP_BLOCK`, so the model's own number is held
against Excel's rather than against our rounded picture.

    python tools\metrics\_xlsx_shape_pitch_size.py
    python tools\metrics\_xlsx_shape_pitch_size.py --reuse
"""
import argparse
import json
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_pitch_size")
BOOK = SCRATCH / "pitchsize.xlsx"
EMU = 9525.0

# How many lines a case sets. The pitch is read as (last top - first top) /
# (LINES - 1), and both ends are rounded to a whole pixel, so the reading is
# worth +-1/(LINES - 1) of a pixel: eight lines resolve a seventh, which is
# not enough to see the 0.06 that puts `tb_r8_jizensoudan`'s third line on
# the wrong side of a boundary. Twenty-four resolve a twenty-third.
LINES = 24
ROW = 20                # the default row, in pixels at 96dpi
# A letter with ink in EVERY row it spans, so a line reads as one run. 国
# fills the em box but is three horizontal bars: at sizes where ClearType
# renders its two thin sides above the threshold, the interior rows go blank
# and one line is read as three (24 "lines" for eight, 10 and 12 point).
WORD = "H"

FACES = ["游ゴシック", "Yu Gothic UI", "メイリオ", "ＭＳ Ｐゴシック"]
SIZES = [9.0, 10.0, 11.0, 12.0, 14.0, 16.0, 18.0, 20.0]


def cases():
    return [(face, size) for face in FACES for size in SIZES]


def bands():
    """Where each case sits, in rows. A case gets the room its own size needs.

    One span for every case would be the largest size's, which at 24 lines of
    20 point is 63 rows — sixty-three rows of nothing under every 9 point
    case. Excel draws a block that overruns its box, so the room has to be
    real: a short band would let one case's last lines fall into the next's.
    """
    held, at = [], 0
    for face, size in cases():
        # Two and a fifth em a line, which is over even メイリオ's 1.3 x 1.34
        # — a band that is too short lets one case's last lines fall into the
        # next's, and the reading there is not wrong so much as meaningless
        # (メイリオ 9pt read a pitch of 33.3 where its own is 23.4).
        tall = int((LINES + 1) * size * 96.0 / 72.0 * 2.2 / ROW) + 2
        held.append((at, tall))
        at += tall
    return held


def anchors_xml():
    held = []
    for index, ((face, points), (row, tall)) in enumerate(zip(cases(), bands())):
        one = (
            f'<a:p><a:pPr algn="l"/>'
            f'<a:r><a:rPr lang="ja-JP" sz="{int(points * 100)}">'
            f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
            f"</a:rPr><a:t>{WORD}</a:t></a:r></a:p>"
        )
        runs = one * LINES
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(200 * EMU)}" cy="{int((tall - 1) * ROW * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f"{runs}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=sum(tall for _, tall in bands()) + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=3600)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "pitchsize.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=3600,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_DUMP_BLOCK="1", OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    # The model's own pitch, in the order the shapes are drawn.
    told = []
    for line in done.stderr.decode("utf-8", "replace").splitlines():
        found = re.search(r"pitch=\[([0-9.]+)", line)
        if found:
            told.append(float(found.group(1)))
    return ours, heights, lane, told


def tops(picture, top, foot, lane):
    """The first row of ink of each line, in the band a case owns."""
    held = (picture[top:foot, lane:] < 128).sum(axis=1)
    found, run = [], None
    for step, lit in enumerate(held):
        if lit:
            if run is None:
                found.append(top + step)
            run = step
        elif run is not None and step - run > 1:
            run = None
    return found


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane, told = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    # No single pitch reproduces Excel's sequence of tops: a fractional
    # START is needed as well, and BOTH have to be solved for at once —
    # holding the pitch at the model's turns two arms empty, which says only
    # that the pair has moved, not which half. Each line asks that
    # `o_i - 0.5 <= f + i x pitch < o_i + 0.5`, so the pair lives in a
    # polygon; this walks the pitch across it and reports how wide it is in
    # each direction, and whether f = 0 — a start on a whole pixel, which is
    # what the renderer assumes — is anywhere inside it.
    def region(held, about):
        walk = [one - held[0] for one in held[1:]]
        pitches, lows, highs, flat = [], [], [], False
        for step in range(-1500, 1501):
            pitch = about + step * 0.0002
            low = max(one - 0.5 - (at + 1) * pitch for at, one in enumerate(walk))
            high = min(one + 0.5 - (at + 1) * pitch for at, one in enumerate(walk))
            if low < high:
                pitches.append(pitch)
                lows.append(low)
                highs.append(high)
                flat = flat or (low <= 0.0 < high)
        if not pitches:
            return None
        return (min(pitches), max(pitches), min(lows), max(highs), flat)

    solved = []
    print(f"{'face':<14}{'pt':>6}{'Excel n':>9}{'Excel pitch':>13}"
          f"{'ours n':>8}{'ours pitch':>12}{'model':>9}{'delta':>9}"
          f"{'pitch range':>16}{'start f':>17}{'f=0':>5}")
    for index, ((face, points), (row, tall)) in enumerate(zip(cases(), bands())):
        top = edges.get(row)
        foot = edges.get(row + tall - 1)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = tops(truth, top, foot, lane)
        ours = tops(mine, top, foot, lane)
        # The slope through ALL the tops, not the two ends. Each end is
        # rounded to a whole pixel, so an end-to-end reading is worth
        # +-1/(LINES - 1) of a pixel — a twenty-third here, which is still
        # four times the 0.05 that is being chased. Least squares over
        # twenty-four of them is worth about a hundredth.
        def walk(held):
            if len(held) < 2:
                return float("nan")
            steps = np.arange(len(held), dtype=float)
            return float(np.polyfit(steps, np.asarray(held, dtype=float), 1)[0])
        model = told[index] if index < len(told) else float("nan")
        gap = walk(ours) - walk(theirs)
        # Which LINE the pixel arrives at, not just how much of one there is
        # over the block: a mean of a seventh reads the same whether the last
        # line alone is out or the drift is spread over all of them.
        drift = ("".join(f"{one - two:+d}" for one, two in zip(ours, theirs))
                 if len(ours) == len(theirs) else "n differs")
        flag = "" if abs(gap) < 0.08 else "  <<"
        found = region(theirs, model) if len(theirs) > 2 else None
        said = ("-", "-", "?") if found is None else (
            f"{found[0]:.3f}-{found[1]:.3f}", f"{found[2]:+.3f},{found[3]:+.3f}",
            "yes" if found[4] else "NO")
        solved.append({"face": face, "points": points, "model": model,
                       "excel_tops": [int(one) for one in theirs],
                       "ours_tops": [int(one) for one in ours],
                       "region": None if found is None else list(found[:4])})
        print(f"{face:<14}{points:>6}{len(theirs):>9}{walk(theirs):>13.3f}"
              f"{len(ours):>8}{walk(ours):>12.3f}{model:>9.3f}{gap:>9.3f}"
              f"{said[0]:>16}{said[1]:>17}{said[2]:>5}"
              f"  {drift}{flag}")
    # The readings, so a model of the start fraction can be tried against
    # them without shooting Excel again.
    (SCRATCH / "solved.json").write_text(json.dumps(solved), encoding="utf-8")


if __name__ == "__main__":
    main()
