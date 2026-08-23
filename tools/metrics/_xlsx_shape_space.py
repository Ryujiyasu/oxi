# -*- coding: utf-8 -*-
r"""Does Excel keep the spaces in a shape's run that does not ask it to?

`sanko_tool`'s panel holds a run of three spaces written `<a:t>   </a:t>`,
with no `xml:space="preserve"` — the spelling OOXML says loses its leading and
trailing whitespace. Twelve pixels is exactly the difference between that
panel's line fitting its box and being broken, so it is worth measuring rather
than assuming.

Each arm is the same words in a box wide enough for them, with the spaces
written one way or the other. Where the ink starts says which spaces survived.

    python tools\metrics\_xlsx_shape_space.py
    python tools\metrics\_xlsx_shape_space.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_space")
BOOK = SCRATCH / "space.xlsx"
EMU = 9525.0

SPACING = 4
FACE, POINTS = "游ゴシック", 12.0
WORD = "国国国"

ARMS = [
    ("bare run", '<a:t>   </a:t>', True),
    ("preserved", '<a:t xml:space="preserve">   </a:t>', True),
    ("in one run", None, False),                    # "   国国国" in a single run
    ("preserved one", None, True),                  # the same, preserved
    ("trailing bare", None, False),                 # "国国国   " then more
    ("no spaces", None, False),
]


def runs_of(index):
    size = int(POINTS * 100)
    dress = f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>'
    rpr = f'<a:rPr lang="ja-JP" sz="{size}">{dress}</a:rPr>'
    if index == 0:      # a bare run of spaces, then the words
        return f"<a:r>{rpr}<a:t>   </a:t></a:r><a:r>{rpr}<a:t>{WORD}</a:t></a:r>"
    if index == 1:      # the same, asking for the spaces to be kept
        return (f'<a:r>{rpr}<a:t xml:space="preserve">   </a:t></a:r>'
                f"<a:r>{rpr}<a:t>{WORD}</a:t></a:r>")
    if index == 2:      # the spaces inside the run that holds the words
        return f"<a:r>{rpr}<a:t>   {WORD}</a:t></a:r>"
    if index == 3:
        return f'<a:r>{rpr}<a:t xml:space="preserve">   {WORD}</a:t></a:r>'
    if index == 4:      # spaces at the end of a run, with a run after it
        return (f"<a:r>{rpr}<a:t>{WORD}   </a:t></a:r>"
                f"<a:r>{rpr}<a:t>{WORD}</a:t></a:r>")
    return f"<a:r>{rpr}<a:t>{WORD}</a:t></a:r>"


def anchors_xml():
    held = []
    for index in range(len(ARMS)):
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(300 * EMU)}" cy="{int(3 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f'<a:p><a:pPr algn="l"/>{runs_of(index)}</a:p></xdr:txBody>'
            f"</xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 50.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(ARMS) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=900)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "space.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=900,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"a space is four pixels at {FACE} {POINTS:g}, so three are twelve")
    print(f"{'arm':<16}{'Excel ink':>12}{'ours':>8}{'width':>8}{'ours':>8}")
    for index, (name, _spelling, _keeps) in enumerate(ARMS):
        top = edges.get(index * SPACING + 1)
        foot = edges.get(index * SPACING + SPACING)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue

        def ink(picture):
            band = (picture[top:foot, lane:] < 128)
            columns = np.flatnonzero(band.any(axis=0))
            if not columns.size:
                return None, None
            return int(columns[0]), int(columns[-1] - columns[0] + 1)

        their_left, their_wide = ink(truth)
        our_left, our_wide = ink(mine)
        mark = "" if (their_left, their_wide) == (our_left, our_wide) else "  <<"
        print(f"{name:<16}{str(their_left):>12}{str(our_left):>8}"
              f"{str(their_wide):>8}{str(our_wide):>8}{mark}")


if __name__ == "__main__":
    main()
