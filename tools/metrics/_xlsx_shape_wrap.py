# -*- coding: utf-8 -*-
r"""How many characters does Excel fit on a line of a shape's text?

With the block model settled (`_xlsx_shape_block.py`), what is left between
`sanko_tool` and Excel is the wrap: our panel takes one line more than Excel's
for the same words in the same box, and a block one line too tall hangs its
first line above a middle-anchored box.

Each case is a shape of an exact width holding one long paragraph. The width
is swept a pixel at a time across several break points, and what is read back
is how wide the first line's ink is — which says how many characters fitted.

    python tools\metrics\_xlsx_shape_wrap.py
    python tools\metrics\_xlsx_shape_wrap.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_wrap")
BOOK = SCRATCH / "wrap.xlsx"
EMU = 9525.0            # per 96-dpi pixel

SPACING = 6             # rows per case
POINTS = 12.0
# A full-width run, and a proportional Latin one: the first says what the box
# is worth in whole characters, the second whether the answer is measured the
# same way when the characters are not all one width.
TEXTS = [("kana", "あ" * 30), ("latin", "Wi " * 12),
         # Marks Japanese typesetting will not start or end a line with, so a
         # break that ignores them lands somewhere else.
         ("kinsoku", "あい、うえお。「かきく」けこさし、すせそ。たちつてと"),
         ("mixed", "１．あい「うえ」、おかき123 abc、くけこさしすせそ")]
FACES = [("ＭＳ Ｐゴシック", 12.0), ("游ゴシック", 12.0)]
WIDTHS = list(range(120, 181))       # pixels, a pixel apart


def cases():
    held = []
    for face, points in FACES:
        for kind, text in TEXTS:
            for width in WIDTHS:
                held.append((face, points, kind, text, width))
    return held


def anchors_xml():
    held = []
    for index, (face, points, _kind, text, width) in enumerate(cases()):
        runs = (f'<a:p><a:pPr algn="l"/><a:r>'
                f'<a:rPr lang="ja-JP" sz="{int(points * 100)}">'
                f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
                f"</a:rPr><a:t>{text}</a:t></a:r></a:p>")
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(width * EMU)}" cy="{int(5 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f"{runs}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "wrap.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def first_line(picture, top, foot, lane):
    """The first line's ink: (width, how many lines the block holds)."""
    held = (picture[top:foot, lane:] < 128)
    rows = held.sum(axis=1)
    lines, run = [], None
    for step, lit in enumerate(rows):
        if lit:
            run = [step, step] if run is None else [run[0], step]
        elif run is not None and step - run[1] > 2:
            lines.append(tuple(run))
            run = None
    if run is not None:
        lines.append(tuple(run))
    if not lines:
        return None, 0
    band = held[lines[0][0]:lines[0][1] + 1]
    columns = np.flatnonzero(band.any(axis=0))
    width = int(columns[-1] - columns[0] + 1) if columns.size else 0
    return width, len(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"{'face':<14}{'text':>9}{'box':>5}"
          f"{'Excel line1':>12}{'lines':>7}{'ours line1':>12}{'lines':>7}"
          f"{'differs':>9}")
    for index, (face, _points, kind, _text, width) in enumerate(cases()):
        top = edges.get(index * SPACING)
        foot = edges.get(index * SPACING + SPACING - 1)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = first_line(truth, top, foot, lane)
        ours = first_line(mine, top, foot, lane)
        # The whole block, where it stands: a break in the wrong place moves
        # every line after it, and the count alone would not see a swap.
        band = (slice(top, foot), slice(lane, min(truth.shape[1], mine.shape[1])))
        differs = int(((truth[band] < 128) ^ (mine[band] < 128)).sum())
        flag = "" if differs == 0 else "  <<"
        print(f"{face:<14}{kind:>9}{width:>5}"
              f"{theirs[0] if theirs[0] is not None else -1:>12}{theirs[1]:>7}"
              f"{ours[0] if ours[0] is not None else -1:>12}{ours[1]:>7}"
              f"{differs:>9}{flag}")


if __name__ == "__main__":
    main()
