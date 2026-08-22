# -*- coding: utf-8 -*-
r"""How far outside its row does Excel let a cell's ink run?

`_xlsx_cramped_profile.py` showed that the ink reading zero at the top of a
cramped row is the row above's, spilling downward with a clear gap under it.
So the question is not how Excel raises a line that does not fit, but where
Excel clips one — the renderer currently cuts at the row's own edges (a pixel
in at the top).

Each case is a short row holding text too tall for it, with a tall empty row
above and below, so any ink outside the row is unambiguous. A second arm puts
text in the neighbouring rows to see whether an occupied neighbour cuts the
bleed short.

    python tools\metrics\_xlsx_row_bleed.py
    python tools\metrics\_xlsx_row_bleed.py --reuse
"""
import argparse
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_row_bleed")
BOOK = SCRATCH / "bleed.xlsx"

CASES = [("ＭＳ Ｐゴシック", 11.0, False, 10.5),
         ("ＭＳ Ｐゴシック", 11.0, False, 6.0),
         ("ＭＳ Ｐゴシック", 18.0, True, 10.5),
         ("ＭＳ Ｐゴシック", 18.0, True, 15.0),
         ("ＭＳ Ｐゴシック", 28.0, False, 12.0),
         ("游ゴシック", 11.0, True, 10.5),
         ("Calibri", 18.0, False, 10.5),
         ("メイリオ", 11.0, False, 9.0)]
PLACES = ["top", "center", "bottom"]
SPACER = 24.0  # points — 32 pixels of clear room either side


def build(neighbours):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 14.0
    sheet.column_dimensions["B"].width = 3.0
    sheet.column_dimensions["C"].width = 14.0
    cases, row = [], 1
    for face, points, bold, height in CASES:
        for place in PLACES:
            for step, filler in ((0, "above"), (2, "below")):
                sheet.row_dimensions[row + step].height = SPACER
                # The filler sits in a column of its own so the case row's
                # own ink can be told from its neighbours' by where it is
                # across the sheet, not only by which band it lands in.
                if neighbours:
                    cell = sheet.cell(row=row + step, column=3, value="あAg")
                    cell.font = Font(name="ＭＳ Ｐゴシック", size=11.0)
                    cell.alignment = Alignment(
                        vertical="bottom" if filler == "above" else "top")
            cell = sheet.cell(row=row + 1, column=1, value="あAg")
            cell.font = Font(name=face, size=points, bold=bold)
            cell.alignment = Alignment(vertical=place, horizontal="left")
            sheet.row_dimensions[row + 1].height = height
            cases.append((row, row + 1, face, points, bold, height, place))
            row += 3
    book.save(BOOK)
    return cases


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def draw():
    ours = SCRATCH / "bleed.oxi.png"
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1",
                       OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(ours), "96"],
                          capture_output=True, timeout=300, env=environment)
    heights, columns = {}, {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
    return ours, heights, columns


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    parser.add_argument("--neighbours", action="store_true",
                        help="put text in the rows above and below")
    args = parser.parse_args()

    cases = build(args.neighbours)
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    ours_png, heights, columns = draw()
    truth = np.asarray(Image.open(picture).convert("L"))
    ours = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = (at, at + heights[index])
        at += heights[index]
    # Only the case column: a filler in another column must not be read as
    # the case row's own ink running past its edge.
    lane = columns.get(min(columns), 0) if columns else 0

    print(f"{'face':<18}{'pt':>5}{'row px':>7}{'place':>8}"
          f"{'Excel above':>12}{'below':>7}{'ours above':>12}{'below':>7}")
    for first, case_row, face, points, bold, height, place in cases:
        if case_row + 1 not in edges:
            continue
        top, foot = edges[case_row]
        window_top, window_foot = edges[first][0], edges[case_row + 1][1]
        window_foot = min(window_foot, truth.shape[0], ours.shape[0])
        if window_foot <= window_top:
            continue

        def reach(image):
            band = (image[window_top:window_foot, :lane] < 128).sum(axis=1)
            lit = np.flatnonzero(band)
            if not lit.size:
                return None
            first_lit = window_top + int(lit[0])
            last_lit = window_top + int(lit[-1])
            return (top - first_lit, last_lit - (foot - 1))

        theirs, mine = reach(truth), reach(ours)
        print(f"{face + (' bold' if bold else ''):<18}{points:>5}{foot - top:>7}{place:>8}"
              f"{(theirs or ('-', '-'))[0]:>12}{(theirs or ('-', '-'))[1]:>7}"
              f"{(mine or ('-', '-'))[0]:>12}{(mine or ('-', '-'))[1]:>7}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
