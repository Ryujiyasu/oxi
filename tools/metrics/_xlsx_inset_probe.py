# -*- coding: utf-8 -*-
"""How far in from the edge of its cell Excel starts, and ends, a line of text.

The vertical question is settled (SX23: the box is the font's own row height
and the baseline is measured per font). This asks the same question sideways:
left-aligned text starts some way in from the left edge, right-aligned text
ends some way in from the right, and a pixel of either shows up on every
column of a data sheet.

Read the way the baseline was: both engines draw the same glyphs at the same
size, so the difference between the two inks is the difference between the two
insets, whatever the rasterizer does at the edges.

    python tools\\metrics\\_xlsx_inset_probe.py
    python tools\\metrics\\_xlsx_inset_probe.py --reuse
"""
import argparse
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
SCRATCH = Path(r"C:\tmp\xlsx_inset")
BOOK = SCRATCH / "inset.xlsx"
TRUTH = SCRATCH / "inset.excel.png"
OURS = SCRATCH / "inset.oxi.png"

FONTS = [("ＭＳ ゴシック", 4.5), ("ＭＳ ゴシック", 6.0), ("ＭＳ ゴシック", 7.5),
         ("ＭＳ ゴシック", 9.0), ("ＭＳ ゴシック", 11.0), ("ＭＳ 明朝", 11.0),
         ("ＭＳ Ｐゴシック", 9.0), ("ＭＳ Ｐゴシック", 11.0), ("ＭＳ Ｐ明朝", 12.0),
         ("Meiryo UI", 11.0), ("游ゴシック", 11.0), ("游明朝", 11.0),
         ("Calibri", 11.0), ("Calibri", 18.0), ("Arial", 11.0),
         ("Times New Roman", 12.0), ("Segoe UI", 11.0)]
# The width of the column, in the characters Excel counts them in.
WIDTHS = [8.0, 12.0, 20.0]
PLACES = ["left", "right", "center"]
# A Latin face has no 亜 of its own, and the face the system stands in with is
# not the one this engine measures with, so the two sides would be compared
# through different fonts. Each face is given letters it owns.
LATIN = {"Calibri", "Arial", "Times New Roman", "Segoe UI", "Consolas"}


def text_for(face):
    return "Hx" if face in LATIN else "亜Ｈ"


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    probes, row = [], 1
    for column, width in enumerate(WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for face, points in FONTS:
        for place in PLACES:
            for column, _ in enumerate(WIDTHS, start=1):
                cell = sheet.cell(row=row, column=column, value=text_for(face))
                cell.font = Font(name=face, size=points)
                cell.alignment = Alignment(vertical="bottom", horizontal=place)
            sheet.row_dimensions[row].height = 30.0
            probes.append((row, face, points, place))
            row += 1
    book.save(BOOK)
    return probes


def shoot():
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{TRUTH.resolve()}", encoding="utf-8")
    TRUTH.unlink(missing_ok=True)
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=300)
    listing.unlink(missing_ok=True)


def draw():
    import os

    OURS.unlink(missing_ok=True)
    environment = dict(os.environ, OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(OURS), "96"],
                          capture_output=True, text=True, encoding="utf-8",
                          errors="replace", timeout=300, env=environment)
    columns = {}
    for line in done.stdout.splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
    return columns


def ink_edges(image, top, bottom, left, right):
    band = image[top:bottom, left:right]
    dark = np.flatnonzero((band < 128).any(axis=0))
    return None if dark.size == 0 else (int(dark[0]), int(dark[-1]))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    probes = build()
    if not args.reuse:
        shoot()
    widths = draw()
    truth = np.asarray(Image.open(TRUTH).convert("L"))
    ours = np.asarray(Image.open(OURS).convert("L"))

    edges, at = [0], 0
    for index in sorted(widths):
        at += widths[index]
        edges.append(at)

    band = 40
    print(f"{'face':<16}{'pt':>5}{'place':>7}"
          + "".join(f"{f'col {width:.0f}: E   O':>16}" for width in WIDTHS))
    disagreed = {}
    for index, (row, face, points, place) in enumerate(probes):
        top = index * band
        line = f"{face:<16}{points:>5.1f}{place:>7}"
        for column in range(len(WIDTHS)):
            left, right = edges[column], edges[column + 1]
            theirs = ink_edges(truth, top, top + band, left, right)
            mine = ink_edges(ours, top, top + band, left, right)
            if theirs is None or mine is None:
                line += f"{'-':>16}"
                continue
            # Left-aligned text is read from its left edge, right-aligned from
            # its right; centred text is read from both, since a pixel of
            # slack can land on either side.
            near = 0 if place == "left" else 1
            gap_in = theirs[near] - (left if near == 0 else right - 1)
            gap_ours = mine[near] - (left if near == 0 else right - 1)
            line += f"{gap_in:>9}{gap_ours:>7}"
            disagreed.setdefault((place, gap_ours - gap_in), []).append(
                (face, points, WIDTHS[column]))
        print(line)

    print()
    print("how far our ink edge sits from Excel's:")
    for place, gap in sorted(disagreed):
        rows = disagreed[(place, gap)]
        print(f"   {place:<7}{gap:+d}px  ×{len(rows):<3} "
              f"{', '.join(f'{face} {points:.0f}' for face, points, _ in rows[:5])}"
              f"{' …' if len(rows) > 5 else ''}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
