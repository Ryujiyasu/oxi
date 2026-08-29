# -*- coding: utf-8 -*-
r"""How much shorter than its pitches is a block, and does it depend on n?

`_xlsx_shape_centre.py` asked one line at six percentages and found the height
Excel centres to be `sum(pitch) + 0.25 x (own - tall)` — the last line keeping
three quarters of the difference between the face's own box and the one the
paragraph asked for. Fifteen arms of eighteen.

Shipping that cost `glossary_05` 0.0329: its two-line boxes came right and a
SEVEN-line block moved a pixel and shattered. Back-solving the two, the
correction has to be in (0.735, 2.735] at two lines and in (-0.93, 1.07] at
seven, and 1.383 — what one line gives — sits in the first and outside the
second. So the term depends on the line count, and one line could not say so.

This sweeps the count directly. For each arm the correction Excel implies is
read straight off the two pictures: a correction `c` moves a centred block up
by `c / 2`, so `c = 2 x (ours - Excel)`.

    python tools\metrics\_xlsx_centre_lines.py
    python tools\metrics\_xlsx_centre_lines.py --reuse
"""
import argparse
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_centre_lines")
BOOK = SCRATCH / "centrelines.xlsx"
EMU = 9525.0
ROW = 20                # a default row, in pixels

WORD = "A"              # sits on the baseline
# (face, size, the face's own line box at that size in pixels). The own box is
# the device line height scaled to the em and multiplied by 1.3 for a Japanese
# face — the pitch a paragraph gets when it asks for nothing.
FACES = [("Yu Gothic UI", 12.0, 27.666), ("メイリオ", 12.0, 31.200)]
# Far from 100%, where `own - tall` is large. The share of it that a block
# loses is read through a one-pixel window, so the window on the SHARE is
# 4 / |own - tall| wide: at 150% that is 0.29 and settles nothing, at 500% it
# is 0.036. This is what pins the sixth that SX124 could only bracket.
PCTS = [80000, 300000]
# Lines that come from WRAPPING one paragraph, against the same number of
# lines written as separate paragraphs. `glossary_05`'s seven-line panel wraps
# and is the one box that refuses the share the sweeps measure, so whether a
# wrapped line counts the same as a written one is the open question.
COUNTS = [1]
WRAPPED = [False]
# Whether the body zeroes its insets. Every arm so far has (to make the slack
# exactly what was asked for); `glossary_05`'s panel keeps Excel's defaults,
# and it is the last structural difference between the arms that show a
# correction and the box that refuses one.
INSETS = [True]
# Whether the run is bold. `glossary_05`'s panel is (`b="1"` on every run) and
# no arm has been — the last attribute separating the two. It does not change
# the face's own line box (measured: Yu Gothic UI and メイリオ give the same
# tmHeight at 400 and 700), so if it changes the correction it does so on its
# own account.
BOLDS = [False, True]
LONG = "A" * 46         # wide enough to wrap in a 200px box
# Whether the body says `vertOverflow="clip"`. Every box the corpus says is
# ALREADY right says it, and no arm of the earlier sweeps did — which is the
# one thing the probe and `glossary_05` did not share.
# Three ways a body can speak about overflow, not two. The arms that implied a
# correction said NOTHING; `glossary_05`'s panel, which takes none, says
# `overflow` out loud. Absent and stated-as-overflow mean the same thing to a
# reader of the schema — so whether they mean the same thing to Excel is the
# question.
CLIPS = ["overflow"]
# The room left over, which IS the slack here: these arms write every inset
# zero, the way `glossary_05`'s little "Yes" and "No" boxes do. The line-count
# sweep gave every arm about 34 pixels of it and found a correction; the boxes
# the corpus says are already right have 1.87 and 8.07. So the slack is the
# axis now.
SPARES = [12]


def cases():
    return [(face, size, own, pct, lines, spare, clip, wrap)
            for face, size, own in FACES for pct in PCTS
            for lines in COUNTS for spare in SPARES for clip in CLIPS
            for wrap in BOLDS]


def bands():
    held, at = [], 0
    for _face, _size, own, pct, lines, spare, _clip, wrap in cases():
        tall = own * pct / 100000.0
        box = int((3 if wrap else lines) * tall + spare)
        rows = box // ROW + 3
        held.append((at, rows, box))
        at += rows
    return held


def anchors_xml():
    held = []
    for index, ((face, points, _own, pct, lines, _spare, clip, wrap), (row, _rows, box)) in enumerate(
            zip(cases(), bands())):
        weight = ' b="1"' if wrap else ""
        clipped = {"absent": "",
                   "overflow": ' vertOverflow="overflow" horzOverflow="overflow"',
                   "clip": ' vertOverflow="clip" horzOverflow="clip"'}[clip]
        one = (
            f'<a:p><a:pPr algn="l">'
            f'<a:lnSpc><a:spcPct val="{pct}"/></a:lnSpc></a:pPr>'
            f'<a:r><a:rPr lang="ja-JP" sz="{int(points * 100)}"{weight}>'
            f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
            f"</a:rPr><a:t>{WORD}</a:t></a:r></a:p>"
        )
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(200 * EMU)}" cy="{int(box * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr{clipped} wrap="square" anchor="ctr"'
            f' lIns="0" tIns="0" rIns="0" bIns="0"/><a:lstStyle/>'
            f"{one * lines}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=sum(one[1] for one in bands()) + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "centrelines.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_DUMP_BLOCK="1", OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    told = []
    for line in done.stderr.decode("utf-8", "replace").splitlines():
        found = re.search(r"^block area (\d+)\.\.(\d+).*?block=([0-9.]+).*? at=([0-9.]+)", line)
        if found:
            told.append(tuple(float(one) for one in found.groups()))
    return ours, heights, lane, told


def first_foot(picture, top, bottom, lane):
    """The foot of the FIRST line of ink — for `A`, a pixel below the baseline."""
    held = (picture[top:bottom, lane:] < 128).sum(axis=1)
    run = None
    for step, lit in enumerate(held):
        if lit:
            run = step
        elif run is not None and step - run > 1:
            return top + run
    return top + run if run is not None else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane, told = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"{'face':<14}{'pct':>6}{'n':>3}{'spare':>7}{'weight':>9}{'own-tall':>10}"
          f"{'Excel':>7}{'ours':>6}{'implied c':>11}{'0.25(o-t)':>11}")
    for index, ((face, _points, own, pct, lines, spare, clip, wrap), (row, rows, _box)) in enumerate(
            zip(cases(), bands())):
        top = edges.get(row)
        bottom = edges.get(row + rows - 1)
        if top is None or bottom is None or index >= len(told):
            continue
        if bottom > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = first_foot(truth, top, bottom, lane)
        ours = first_foot(mine, top, bottom, lane)
        if theirs is None or ours is None:
            continue
        tall = own * pct / 100000.0
        # A correction c makes the block taller and so lifts a centred block
        # by c / 2. Read the other way: c = 2 x (ours - Excel).
        print(f"{face:<14}{pct // 1000:>6}{lines:>3}{spare:>7}{("bold" if wrap else "plain"):>9}"
              f"{own - tall:>10.3f}{theirs:>7}{ours:>6}"
              f"{2 * (ours - theirs):>11}{0.25 * (own - tall):>11.3f}")


if __name__ == "__main__":
    main()
