# -*- coding: utf-8 -*-
"""Where does Excel put the button of a filtered column, and how big is it?

A sheet-level `<autoFilter>` puts a dropdown in every heading cell of its
range — the procurement lists have one on a 90px heading row — and the
renderer only draws buttons for a *table's* filter. This asks Excel for the
button's box against rows of several heights and columns of several widths.

    python tools\\metrics\\_xlsx_filter_button.py
    python tools\\metrics\\_xlsx_filter_button.py --reuse
"""
import argparse
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_filter")
BOOK = SCRATCH / "filter.xlsx"
TRUTH = SCRATCH / "filter.excel.png"

HEIGHTS = [20, 30, 45, 90]          # pixels for the heading row
WIDTHS = [8.0, 14.0]                # characters


def build():
    """One workbook per heading height: a sheet keeps only one filter."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.filters import AutoFilter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    held = []
    for height in HEIGHTS:
        book = Workbook()
        sheet = book.active
        for column, width in enumerate(WIDTHS, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
            cell = sheet.cell(row=1, column=column, value="見出し")
            cell.font = Font(name="ＭＳ Ｐゴシック", size=11)
        sheet.row_dimensions[1].height = height * 0.75
        sheet.cell(row=2, column=1, value="あ")
        sheet.row_dimensions[2].height = 20 * 0.75
        sheet.auto_filter = AutoFilter(ref="A1:B2")
        path = SCRATCH / f"filter_{height}.xlsx"
        book.save(path)
        held.append((height, path))
    return held


def shoot(books):
    listing = SCRATCH / "_batch.txt"
    lines = []
    for _, path in books:
        picture = path.with_suffix(".excel.png")
        picture.unlink(missing_ok=True)
        lines.append(f"{path.resolve()}\t{picture.resolve()}")
    listing.write_text("\n".join(lines), encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=300)
    listing.unlink(missing_ok=True)


def columns(path):
    import os

    environment = dict(os.environ, OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(path), str(SCRATCH / "filter.oxi.png"), "96"],
                          capture_output=True, timeout=300, env=environment)
    held = {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "column":
            held[int(parts[1])] = int(float(parts[3]))
    return held


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    books = build()
    if not args.reuse:
        shoot(books)
    print(f"{'row px':>7}{'column':>8}{'button across the cell':>26}{'and down it':>18}")
    for height, path in books:
        widths = columns(path)
        edges, at = [0], 0
        for index in sorted(widths):
            at += widths[index]
            edges.append(at)
        picture = path.with_suffix(".excel.png")
        if not picture.exists():
            print(f"{height:>7}   (Excel gave no picture)")
            continue
        truth = np.asarray(Image.open(picture).convert("L"))
        for column in range(len(WIDTHS)):
            left, right = edges[column], edges[column + 1]
            band = truth[:height, left:right]
            grey = (band > 120) & (band < 220)
            across = np.flatnonzero(grey.sum(axis=0) >= 3)
            down = np.flatnonzero(grey.sum(axis=1) >= 3)
            if across.size == 0 or down.size == 0:
                print(f"{height:>7}{column:>8}{'(no button)':>26}")
                continue
            print(f"{height:>7}{column:>8}"
                  f"{f'{across[0]}..{across[-1]} of {right - left}':>26}"
                  f"{f'{down[0]}..{down[-1]} of {height}':>18}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
