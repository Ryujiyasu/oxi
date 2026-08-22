# -*- coding: utf-8 -*-
"""Which pixels along a broken rule does Excel ink?

The renderer inks a broken rule run by run from the cell's own edge, so the
pattern restarts at every cell and its phase follows the cell. Excel's hair
rules in the corpus look like a halftone anchored somewhere else entirely:
on `data_B28` every hair rule is dark exactly where `(x + y)` is even, which
no per-cell run can produce.

This asks Excel directly. One workbook per border style, a grid of boxes at
deliberately mixed odd and even column and row boundaries, and the ink along
each rule read straight off Excel's own picture.

    python tools\\metrics\\_xlsx_border_pattern.py
    python tools\\metrics\\_xlsx_border_pattern.py --reuse
"""
import argparse
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_border")

STYLES = ["hair", "dotted", "dashed", "dashDot", "dashDotDot", "thin", "medium", "mediumDashed"]
WIDTHS = [3.0, 4.29, 5.57, 6.86]        # characters: boundaries land on mixed parities
HEIGHTS = [15.0, 15.75, 18.0, 21.75]    # points


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    SCRATCH.mkdir(parents=True, exist_ok=True)
    held = []
    for style in STYLES:
        book = Workbook()
        sheet = book.active
        side = Side(style=style, color="FF000000")
        for column, width in enumerate(WIDTHS, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        for row, height in enumerate(HEIGHTS, start=1):
            sheet.row_dimensions[row].height = height
            for column in range(1, len(WIDTHS) + 1):
                sheet.cell(row=row, column=column).border = Border(
                    left=side, right=side, top=side, bottom=side)
        path = SCRATCH / f"border_{style}.xlsx"
        book.save(path)
        held.append((style, path))
    return held


def shoot(books):
    listing = SCRATCH / "_batch.txt"
    lines = []
    for _, path in books:
        picture = path.with_suffix(".excel.png")
        picture.unlink(missing_ok=True)
        lines.append(f"{path.resolve()}\t{picture.resolve()}")
    listing.write_text("\n".join(lines), encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)


def geometry(path):
    """The renderer's own column widths and row heights, to find the rules."""
    import os

    environment = dict(os.environ, OXI_XLSX_DUMP_COLUMNS="1", OXI_XLSX_DUMP_ROWS="1")
    done = subprocess.run([str(RENDERER), str(path), str(SCRATCH / "border.oxi.png"), "96"],
                          capture_output=True, timeout=300, env=environment)
    columns, rows = {}, {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
        elif len(parts) == 4 and parts[0] == "row":
            rows[int(parts[1])] = int(float(parts[3]))
    def edges(sizes):
        out, at = [0], 0
        for index in sorted(sizes):
            at += sizes[index]
            out.append(at)
        return out
    return edges(columns), edges(rows)


def ink(strip):
    return "".join("#" if v < 128 else "." for v in strip)


def parity(dark, first):
    """Are the inked pixels all at an even (or all at an odd) sum with `first`?"""
    sums = {(first + i) % 2 for i, v in enumerate(dark) if v}
    return {0: "even", 1: "odd"}.get(next(iter(sums))) if len(sums) == 1 else "mixed"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    books = build()
    if not args.reuse:
        shoot(books)

    for style, path in books:
        picture = path.with_suffix(".excel.png")
        if not picture.exists():
            print(f"{style}: (Excel gave no picture)")
            continue
        truth = np.asarray(Image.open(picture).convert("L"))
        across, down = geometry(path)
        print(f"\n=== {style} === picture {truth.shape[1]}x{truth.shape[0]},"
              f" columns at {across}, rows at {down}")
        for y in down[:-1]:
            if y >= truth.shape[0]:
                continue
            left, right = across[0], min(across[-1], truth.shape[1])
            strip = truth[y, left:right] < 128
            print(f"  rule along y={y:<4} x+y {parity(strip, left + y):<5} {ink(truth[y, left:right])}")
        for x in across[:-1]:
            if x >= truth.shape[1]:
                continue
            top, bottom = down[0], min(down[-1], truth.shape[0])
            strip = truth[top:bottom, x] < 128
            print(f"  rule along x={x:<4} x+y {parity(strip, top + x):<5} {ink(truth[top:bottom, x])}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
