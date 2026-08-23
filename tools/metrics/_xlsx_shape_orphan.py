# -*- coding: utf-8 -*-
r"""Will Excel leave one character alone on a shape's last line?

`sanko_tool`'s panel breaks its line at 「…選択」and puts 「する」on the next,
where this renderer fits one character more and leaves 「る」 by itself. Both
breaks are inside the room, so what is being tested is not the room: it is
whether Excel holds a character back rather than end a paragraph with a line
of one.

Each case is a run of full-width characters in a box whose room is a whole
number of them, so the leftover is known: a room of sixteen characters for a
run of seventeen leaves one, for eighteen leaves two, and so on. What is read
back is how wide the last line's ink is.

    python tools\metrics\_xlsx_shape_orphan.py
    python tools\metrics\_xlsx_shape_orphan.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_orphan")
BOOK = SCRATCH / "orphan.xlsx"
EMU = 9525.0

SPACING = 8
FACE, POINTS = "游ゴシック", 12.0
EM = 16.0                       # one full-width character at this size
INSET = 2 * 91440 / EMU
FITS = 12                       # characters a line has room for
# A run of this many characters, so the last line would hold one, two, three
# — and the same again with a Latin tail, where a word may not be split.
LENGTHS = [13, 14, 15, 16, 24, 25, 26]
KINDS = [("kana", "\u3042"), ("kanji", "\u56fd")]


def cases():
    return [(kind, letter, length) for kind, letter in KINDS for length in LENGTHS]


def anchors_xml():
    held = []
    for index, (_kind, letter, length) in enumerate(cases()):
        size = int(POINTS * 100)
        runs = (f'<a:p><a:pPr algn="l"/><a:r>'
                f'<a:rPr lang="ja-JP" sz="{size}">'
                f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>'
                f"</a:rPr><a:t>{letter * length}</a:t></a:r></a:p>")
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{round((FITS * EM + INSET) * EMU)}" cy="{int(7 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f"{runs}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=900)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "orphan.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=900,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def lines_of(picture, top, foot, lane):
    """Each line's ink width, in order."""
    band = (picture[top:foot, lane:] < 128)
    rows = band.sum(axis=1)
    held, run = [], None
    for step, lit in enumerate(rows):
        if lit:
            run = [step, step] if run is None else [run[0], step]
        elif run is not None and step - run[1] > 2:
            held.append(run)
            run = None
    if run is not None:
        held.append(run)
    widths = []
    for first, last in held:
        columns = np.flatnonzero(band[first:last + 1].any(axis=0))
        widths.append(int(columns[-1] - columns[0] + 1) if columns.size else 0)
    return widths


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"the room holds {FITS} characters of {EM:g} pixels")
    print(f"{'text':>7}{'length':>7}{'Excel lines':>28}{'ours':>28}")
    for index, (kind, _letter, length) in enumerate(cases()):
        top = edges.get(index * SPACING + 1)
        foot = edges.get(index * SPACING + SPACING)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = lines_of(truth, top, foot, lane)
        ours = lines_of(mine, top, foot, lane)
        mark = "" if theirs == ours else "  <<"
        print(f"{kind:>7}{length:>7}{str(theirs):>28}{str(ours):>28}{mark}")


if __name__ == "__main__":
    main()
