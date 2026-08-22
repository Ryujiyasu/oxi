# -*- coding: utf-8 -*-
"""Where does Excel put a line inside its row — measured in the gate's pixels?

The earlier `_xlsx_valign_probe.py` asked a PDF, which carries a scale of its
own and cannot answer a question worth one pixel. This asks Excel's own
picture instead: a sheet of one-line cells varying only in font, size, row
height and vertical alignment, shot with `CopyPicture`, drawn by the renderer
at the same size, and the ink compared band by band.

The `h2daa*kre` family sits a pixel low in exactly its vertically centred
rows, so what is being tested is which way Excel rounds the odd pixel of the
slack, and whether the answer depends on the font or on the slack itself.

    python tools\\metrics\\_xlsx_valign_pixels.py
    python tools\\metrics\\_xlsx_valign_pixels.py --reuse
"""
import argparse
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_valign_px")
BOOK = SCRATCH / "valign.xlsx"
MERGED = SCRATCH / "valign_merged.xlsx"

FONTS = [("ＭＳ Ｐゴシック", 11.0), ("ＭＳ Ｐゴシック", 12.0), ("ＭＳ Ｐゴシック", 14.0),
         ("ＭＳ ゴシック", 11.0), ("ＭＳ 明朝", 11.0), ("Meiryo UI", 11.0),
         ("游ゴシック", 11.0), ("Calibri", 11.0)]
# Points. Each is a different slack against the font's own line box, so the
# odd pixel lands both ways round.
HEIGHTS = [10.5, 11.25, 12.0, 12.75, 13.5, 14.25, 15.0, 15.75, 17.25, 18.0,
           20.25, 24.0, 30.0]
PLACES = ["top", "center", "bottom"]


def build(merged=False, deep=1, lines=1):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    for column in "ABC":
        sheet.column_dimensions[column].width = 12.0
    cases, row = [], 1
    for face, points in FONTS:
        for height in HEIGHTS:
            for place in PLACES:
                cell = sheet.cell(row=row, column=1,
                                  value=chr(10).join(["あA"] * lines))
                cell.font = Font(name=face, size=points)
                cell.alignment = Alignment(vertical=place, horizontal="left",
                                           wrap_text=lines > 1)
                for step in range(deep):
                    sheet.row_dimensions[row + step].height = height * lines / deep
                # The `h2daa*kre` rows that sit a pixel low are all merged
                # across their columns, and the ones that do not are not.
                if merged or deep > 1:
                    sheet.merge_cells(start_row=row, start_column=1,
                                      end_row=row + deep - 1, end_column=3)
                cases.append((row, face, points, height, place, deep))
                row += deep
    book.save(BOOK)
    return cases


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def draw():
    ours = SCRATCH / "valign.oxi.png"
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(ours), "96"],
                          capture_output=True, timeout=300, env=environment)
    heights = {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
    return ours, heights


def ink_top(image, top, foot):
    band = (image[top:foot] < 128).sum(axis=1)
    lit = np.flatnonzero(band)
    return (int(lit[0]), int(lit[-1])) if lit.size else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    parser.add_argument("--merged", action="store_true",
                        help="merge every case across its columns")
    parser.add_argument("--deep", type=int, default=1,
                        help="merge every case down this many rows as well")
    parser.add_argument("--lines", type=int, default=1,
                        help="how many lines of text each cell holds")
    args = parser.parse_args()

    cases = build(args.merged, args.deep, args.lines)
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    ours_png, heights = draw()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours = np.asarray(Image.open(ours_png).convert("L"))

    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = (at, at + heights[index])
        at += heights[index]

    print(f"{'font':<16}{'pt':>5}{'ht':>7}{'px':>4}{'place':>8}"
          f"{'Excel':>8}{'ours':>7}{'delta':>7}"
          f"{'Excel':>8}{'ours':>7}{'delta':>7}   (ink top, then ink foot)")
    trouble = 0
    for row, face, points, height, place, deep in cases:
        if row not in edges or (row + deep - 1) not in edges:
            continue
        top, foot = edges[row][0], edges[row + deep - 1][1]
        if foot > min(truth.shape[0], ours.shape[0]):
            continue
        theirs = ink_top(truth, top, foot)
        mine = ink_top(ours, top, foot)
        if theirs is None or mine is None:
            print(f"{face:<16}{points:>5}{height:>7}{foot - top:>4}{place:>8}"
                  f"{'(no ink)' if theirs is None else '':>8}{'(no ink)' if mine is None else '':>7}")
            continue
        delta = mine[0] - theirs[0]
        below = mine[1] - theirs[1]
        trouble += delta != 0 or below != 0
        print(f"{face:<16}{points:>5}{height:>7}{foot - top:>4}{place:>8}"
              f"{theirs[0]:>8}{mine[0]:>7}{delta:>+7}"
              f"{theirs[1]:>8}{mine[1]:>7}{below:>+7}")
    print(f"\n{trouble} of {len(cases)} rows sit at a different height")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
