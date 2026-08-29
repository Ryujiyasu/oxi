# -*- coding: utf-8 -*-
r"""Where does the text of a group's child go when the group is stretched?

`glossary_05`'s flowchart is four groups, and the shapes still a pixel or two
out are all children of one that stretches 1.15x down. Their boxes agree with
Excel — fill and outline land exactly — and only the text is low. The implied
Excel position wants the box's FOOT about four pixels higher than the drawn
one, which is what a text frame that ignored the stretch would look like.

So: a group whose child is a fixed rectangle in child space, drawn at several
vertical scales, with the text anchored top in one lane and centred in another.
Reading the child's own box off the picture and the text against it separates
the two stories —

  * laid out in the box as DRAWN: the gap above centred text grows with the
    scale;
  * laid out in the child's own unstretched box: it does not.

    python tools\metrics\_xlsx_group_scale_text.py
    python tools\metrics\_xlsx_group_scale_text.py --reuse
"""
import argparse
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_group_scale_text")
BOOK = SCRATCH / "groupscale.xlsx"
EMU = 9525.0

SPACING = 12            # rows a case
WORD = "A"              # sits on the baseline
FACE, POINTS = "Yu Gothic UI", 12.0
CHILD = (150.0, 60.0)   # the child's size in the group's own coordinates, px
SCALES = [1.0, 1.15, 1.5, 2.0, 2.5]
ANCHORS = ["t", "ctr"]


def cases():
    return [(scale, anchor) for scale in SCALES for anchor in ANCHORS]


def anchors_xml():
    held = []
    for index, (scale, anchor) in enumerate(cases()):
        wide, tall = CHILD
        # The group states the room it occupies (ext) and the coordinates its
        # children are written in (chExt). Their ratio is the stretch.
        ext = (int(wide * EMU), int(tall * scale * EMU))
        ch_ext = (int(wide * EMU), int(tall * EMU))
        one = (
            f'<a:p><a:pPr algn="l"/>'
            f'<a:r><a:rPr lang="ja-JP" sz="{int(POINTS * 100)}">'
            f'<a:latin typeface="{FACE}"/><a:ea typeface="{FACE}"/>'
            f"</a:rPr><a:t>{WORD}</a:t></a:r></a:p>"
        )
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{ext[0]}" cy="{ext[1]}"/>'
            f"<xdr:grpSp>"
            f'<xdr:nvGrpSpPr><xdr:cNvPr id="{index * 2 + 2}" name="group {index}"/>'
            f"<xdr:cNvGrpSpPr/></xdr:nvGrpSpPr>"
            f"<xdr:grpSpPr><a:xfrm>"
            f'<a:off x="0" y="0"/><a:ext cx="{ext[0]}" cy="{ext[1]}"/>'
            f'<a:chOff x="0" y="0"/><a:chExt cx="{ch_ext[0]}" cy="{ch_ext[1]}"/>'
            f"</a:xfrm></xdr:grpSpPr>"
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index * 2 + 3}" name="child {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/>'
            f'<a:ext cx="{ch_ext[0]}" cy="{ch_ext[1]}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            # Filled, so the box the child is actually drawn in is readable.
            f'<a:solidFill><a:srgbClr val="D9D9D9"/></a:solidFill>'
            f"<a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="{anchor}"/><a:lstStyle/>'
            f"{one}</xdr:txBody></xdr:sp>"
            f"</xdr:grpSp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "groupscale.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def read(picture, top, bottom, lane):
    """The child's own box (its fill) and the foot of the ink inside it."""
    band = picture[top:bottom, lane + 4:lane + 140]
    fill = np.abs(band.astype(int) - 217) < 6
    rows = np.where(fill.sum(axis=1) > 40)[0]
    if len(rows) == 0:
        return None
    box_top, box_foot = top + int(rows[0]), top + int(rows[-1])
    ink = (band < 128).sum(axis=1)
    lit = np.where(ink > 0)[0]
    if len(lit) == 0:
        return box_top, box_foot, None
    return box_top, box_foot, top + int(lit[-1])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"child {CHILD[0]:.0f}x{CHILD[1]:.0f}px in the group's own coordinates, "
          f"{FACE} {POINTS}pt")
    print(f"{'scale':>6}{'anchor':>8}   {'Excel box/foot/gap':>28}{'ours box/foot/gap':>28}")
    for index, (scale, anchor) in enumerate(cases()):
        top = edges.get(index * SPACING)
        bottom = edges.get(index * SPACING + SPACING - 1)
        if top is None or bottom is None:
            continue
        if bottom > min(truth.shape[0], mine.shape[0]):
            continue
        held = []
        for dark in (truth, mine):
            found = read(dark, top, bottom, lane)
            if found is None or found[2] is None:
                held.append("-")
                continue
            box_top, box_foot, ink = found
            # The gap above the ink is what tells the two stories apart.
            held.append(f"{box_foot - box_top + 1}px {ink - box_top} {box_foot - ink}")
        same = held[0] == held[1]
        print(f"{scale:>6.2f}{anchor:>8}   {held[0]:>28}{held[1]:>28}"
              f"{'' if same else '  <<'}")


if __name__ == "__main__":
    main()
