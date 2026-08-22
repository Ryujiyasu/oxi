# -*- coding: utf-8 -*-
r"""How short must a row be before Excel cuts the ink off at its edge?

`_xlsx_bleed_pair.py` found that ＭＳ Ｐゴシック at 11 point runs its ink past
the row's foot in a 16- and 17-pixel row and is cut dead at the row's foot in
a 15-pixel one, whatever the row below holds. So there is a threshold, and it
is worth knowing what it is made of: the line box, the baseline, the device's
tmHeight, its ascent, or the em.

Each case is one row of one height holding one line, with a tall empty spacer
above and below so ink either side of the row is unambiguous, and the text in
a column of its own. Both ends are reported: how far above the row's top the
ink starts and how far below its foot it ends.

    python tools\metrics\_xlsx_bleed_threshold.py
    python tools\metrics\_xlsx_bleed_threshold.py --reuse
"""
import argparse
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_bleed_threshold")
BOOK = SCRATCH / "threshold.xlsx"
WIDE_BOOK = SCRATCH / "threshold_wide.xlsx"

FACES = [("ＭＳ Ｐゴシック", 11.0, False), ("ＭＳ 明朝", 11.0, False),
         ("ＭＳ Ｐゴシック", 18.0, True), ("游ゴシック", 11.0, False),
         ("游ゴシック", 11.0, True), ("Calibri", 11.0, False),
         ("メイリオ", 11.0, False), ("Arial", 10.0, False),
         ("ＭＳ Ｐゴシック", 14.0, False), ("BIZ UDPゴシック", 11.0, False)]
# A second, wider arm: the same question over enough faces and sizes for a
# formula over the font's own numbers to be told from a table of measurements.
WIDE = [("ＭＳ Ｐゴシック", 18.0, False), ("ＭＳ Ｐゴシック", 11.0, True),
        ("ＭＳ Ｐゴシック", 14.0, True), ("ＭＳ ゴシック", 11.0, False),
        ("ＭＳ ゴシック", 18.0, False), ("游ゴシック", 14.0, False),
        ("游ゴシック", 18.0, False), ("Calibri", 14.0, False),
        ("Calibri", 18.0, False), ("メイリオ", 14.0, False),
        ("Arial", 11.0, False), ("Arial", 18.0, False),
        ("Meiryo UI", 11.0, False), ("BIZ UDPゴシック", 18.0, False),
        ("Yu Gothic UI", 11.0, False), ("ＭＳ 明朝", 18.0, False)]
# Points, a pixel apart: 0.75pt is one pixel at 96 dpi.
STEPS = [6.0 + 0.75 * step for step in range(28)]      # 8 … 35 pixels
WIDE_STEPS = [6.0 + 0.75 * step for step in range(40)]  # 8 … 47 pixels
PLACES = ["top", "bottom"]
TEXT = "あAg"
SPACER = 27.0


def build(wide=False, lines=1, wrap=False, number=False, merged=False):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 12.0
    cases, row = [], 1
    faces, steps = (WIDE, WIDE_STEPS) if wide else (FACES, STEPS)
    places = ["top"] if wide else PLACES
    for face, points, bold in faces:
        for height in steps:
            for place in places:
                sheet.row_dimensions[row].height = SPACER
                cell = sheet.cell(row=row + 1, column=1,
                                  value=1234567 if number
                                  else chr(10).join([TEXT] * lines))
                cell.font = Font(name=face, size=points, bold=bold)
                cell.alignment = Alignment(vertical=place, horizontal="left",
                                           wrap_text=wrap or lines > 1)
                sheet.row_dimensions[row + 1].height = height
                sheet.row_dimensions[row + 2].height = SPACER
                # A merged block is one box across two rows: the same
                # question asked of a box that is not a row.
                if merged:
                    sheet.merge_cells(start_row=row + 1, start_column=1,
                                      end_row=row + 1, end_column=2)
                cases.append((row + 1, face, points, bold, place))
                row += 3
    book.save(WIDE_BOOK if wide else BOOK)
    return cases


def shoot(book):
    picture = book.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{book.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=900)
    listing.unlink(missing_ok=True)
    return picture


def draw(book):
    ours = book.with_suffix(".oxi.png")
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(book), str(ours), "96"],
                          capture_output=True, timeout=600, env=environment)
    heights, columns = {}, {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
    return ours, heights, columns


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    parser.add_argument("--wide", action="store_true",
                        help="more faces and sizes, top-aligned only")
    parser.add_argument("--lines", type=int, default=1,
                        help="how many wrapped lines the case cell holds")
    parser.add_argument("--wrap", action="store_true",
                        help="turn wrapping on even for a single line")
    parser.add_argument("--number", action="store_true",
                        help="hold a number rather than text")
    parser.add_argument("--merged", action="store_true",
                        help="merge the case cell across two columns")
    args = parser.parse_args()

    cases = build(args.wide, args.lines, args.wrap, args.number, args.merged)
    book = WIDE_BOOK if args.wide else BOOK
    picture = book.with_suffix(".excel.png") if args.reuse else shoot(book)
    ours_png, heights, columns = draw(book)
    truth = np.asarray(Image.open(picture).convert("L"))
    ours = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = (at, at + heights[index])
        at += heights[index]
    lane = columns.get(min(columns), 0) if columns else 0

    print(f"{'face':<18}{'pt':>5}{'row':>5}{'place':>8}"
          f"{'Excel head':>11}{'foot':>6}{'ours head':>11}{'foot':>6}"
          f"   (pixels outside the row, + is outside)")
    for case_row, face, points, bold, place in cases:
        if case_row + 1 not in edges or case_row - 1 not in edges:
            continue
        top, foot = edges[case_row]
        start = edges[case_row - 1][0]
        stop = min(edges[case_row + 1][1], truth.shape[0], ours.shape[0])
        if stop <= start:
            continue

        def reach(image):
            band = (image[start:stop, :lane] < 128).sum(axis=1)
            lit = np.flatnonzero(band)
            if not lit.size:
                return None, None
            return top - (start + int(lit[0])), (start + int(lit[-1])) - (foot - 1)

        their_head, their_foot = reach(truth)
        our_head, our_foot = reach(ours)
        flag = "" if (their_head, their_foot) == (our_head, our_foot) else "  <<"
        print(f"{face + (' bold' if bold else ''):<18}{points:>5}{foot - top:>5}{place:>8}"
              f"{str(their_head):>11}{str(their_foot):>6}"
              f"{str(our_head):>11}{str(our_foot):>6}{flag}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
