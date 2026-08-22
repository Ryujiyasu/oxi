# -*- coding: utf-8 -*-
r"""Where does Excel put each character of a shape's line?

`sanko_tool`'s panel takes one line more than Excel's for the same words, and
`_xlsx_shape_wrap.py` narrowed that to the run of glyphs rather than the box:
the line counts agree for kana and part company for proportional Latin, and
positions inside a line differ by a pixel or two, more the longer the line.

This asks the question one character at a time. Each case is a shape holding
the first N characters of a string, set not to wrap, so the ink's right edge is
where the run has got to by character N. The first N at which Excel and the
renderer part company names the character whose advance is wrong.

    python tools\metrics\_xlsx_shape_advance.py
    python tools\metrics\_xlsx_shape_advance.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_advance")
BOOK = SCRATCH / "advance.xlsx"
EMU = 9525.0

SPACING = 3             # rows per case: one line each
LONGEST = 26
# The workbook's own words, the kana that wrap the same way, and a
# proportional Latin run that does not.
STRINGS = [
    # One character repeated: the ink either end is then the same ink, so the
    # difference between two lengths is the advance itself and nothing else.
    ("one ア", "ア" * 30),
    ("one W", "W" * 30),
    ("one i", "i" * 30),
]
FONTS = [("游ゴシック", 12.0), ("游ゴシック", 11.0), ("ＭＳ Ｐゴシック", 12.0),
         ("ＭＳ Ｐゴシック", 11.0), ("ＭＳ 明朝", 11.0), ("メイリオ", 11.0),
         ("メイリオ", 12.0), ("Calibri", 11.0), ("Calibri", 12.0),
         ("Meiryo UI", 11.0)]


def cases():
    return [(face, points, kind, text[:count])
            for face, points in FONTS
            for kind, text in STRINGS
            for count in range(1, LONGEST + 1)]


def anchors_xml():
    held = []
    for index, (face, points, _kind, text) in enumerate(cases()):
        # `&` and `<` never appear in these strings, so the text goes in whole.
        runs = (f'<a:p><a:pPr algn="l"/><a:r>'
                f'<a:rPr lang="ja-JP" sz="{int(points * 100)}">'
                f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
                f"</a:rPr><a:t>{text}</a:t></a:r></a:p>")
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(900 * EMU)}" cy="{int(2 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="none" anchor="t"/><a:lstStyle/>'
            f"{runs}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 120.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "advance.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def edges(picture, top, foot, lane):
    band = (picture[top:foot, lane:] < 128)
    columns = np.flatnonzero(band.any(axis=0))
    if not columns.size:
        return None, None
    return int(columns[0]), int(columns[-1])


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    at, starts = 0, {}
    for index in sorted(heights):
        starts[index] = at
        at += heights[index]

    print(f"{'face':<14}{'pt':>5}{'text':>7}{'n':>4}"
          f"{'Excel right':>12}{'ours right':>11}{'d':>4}   last character")
    for index, (face, points, kind, text) in enumerate(cases()):
        top = starts.get(index * SPACING + 1)
        foot = starts.get(index * SPACING + SPACING)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        _, theirs = edges(truth, top, foot, lane)
        _, ours = edges(mine, top, foot, lane)
        if theirs is None or ours is None:
            continue
        mark = "" if theirs == ours else "  <<"
        print(f"{face:<14}{points:>5}{kind:>7}{len(text):>4}"
              f"{theirs:>12}{ours:>11}{ours - theirs:>+4}   {text[-1]!r}{mark}")


if __name__ == "__main__":
    main()
