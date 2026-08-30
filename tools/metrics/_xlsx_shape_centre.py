# -*- coding: utf-8 -*-
r"""How does Excel divide what is left over when a block is centred?

The renderer puts a centred block at `area top + round(slack / 2)`, with
`slack` the room the box has left after the block's own (fractional) height.
`glossary_05`'s flowchart is eight centred boxes out of twelve and four of
them still sit a pixel or two from Excel's, so the division is worth asking
about rather than assuming.

The area is a whole-pixel rectangle, so the slack moves a pixel at a time and
half of it alternates between a whole number and a half. Sweeping the box's
height one pixel at a time therefore walks the rounding boundary again and
again: over sixteen heights the three candidate rules give three different
sequences.

Three faces at one size, whose own line heights end in .666, .200 and .800, so
the sweep meets the boundary in a different phase in each.

    python tools\metrics\_xlsx_shape_centre.py
    python tools\metrics\_xlsx_shape_centre.py --reuse
"""
import argparse
import os
import re
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_centre")
BOOK = SCRATCH / "centre.xlsx"
EMU = 9525.0

# Rows between one case and the next. A row is about twenty pixels, so this has
# to clear the TALLEST box in `cases()` — the seven-line arms are 200 pixels and
# at 150% their block overflows the box, which then spills into the arm below
# and the reader answers with its neighbour's ink.
SPACING = 12            # rows a case
WORD = "A"              # sits on the baseline, so its ink foot is baseline - 1
FACES = [("Yu Gothic UI", 12.0), ("メイリオ", 12.0), ("ＭＳ Ｐゴシック", 12.0)]
HEIGHTS = list(range(40, 56))
# A second sweep, on a question the first cannot ask. `glossary_05` sets its
# flowchart at 80% and centres it, and the height Excel CENTRES may not be the
# height it spaces lines by. For one line the two stories separate cleanly:
# centring the scaled pitch moves the baseline by a quarter of the face's own
# line per unit of percentage, centring the face's own line moves it by three
# quarters. Six percentages at one box height.
PCTS = [70000, 80000, 90000, 100000, 115000, 150000]
PCT_HEIGHT = 74         # the height `glossary_05`'s own centred box works out to


# And a third sweep, on the question the second could not answer. The line
# spacing that came out of the one-line arms — Excel divides
# `own × (0.25 + 0.75p)` rather than the sum of the pitches — was implemented
# and cost `glossary_05` 0.0329: the two-line boxes came right and the
# seven-line ones broke, so the correction depends on how many lines there
# are. These arms hold the box and vary the count.
LINES = [1, 2, 3, 7]
LINE_HEIGHT = 200       # tall enough for seven lines at 150%


# A staircase: the box height swept a pixel at a time AT a given percentage.
# Reading the delta on its own only ever answers in whole pixels, and the term
# being looked for is under two of them across the whole range of percentages.
# Walking the height instead moves the rounding boundary past the answer over
# and over, and WHERE the staircase steps says what height Excel is halving to
# sub-pixel precision. Set by `--staircase`.
STAIRS: list[tuple[str, float, int, int | None, int]] = []


def cases():
    if STAIRS:
        return STAIRS
    plain = [(face, size, tall, None, 1) for face, size in FACES for tall in HEIGHTS]
    scaled = [(face, size, PCT_HEIGHT, pct, 1) for face, size in FACES for pct in PCTS]
    counted = [(face, size, LINE_HEIGHT, pct, lines)
               for face, size in FACES for pct in PCTS for lines in LINES]
    return plain + scaled + counted


def anchors_xml():
    held = []
    for index, (face, points, tall, pct, lines) in enumerate(cases()):
        spec = "" if pct is None else f'<a:lnSpc><a:spcPct val="{pct}"/></a:lnSpc>'
        one = (
            f'<a:p><a:pPr algn="l">{spec}</a:pPr>'
            f'<a:r><a:rPr lang="ja-JP" sz="{int(points * 100)}">'
            f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
            f"</a:rPr><a:t>{WORD}</a:t></a:r></a:p>"
        ) * lines
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(200 * EMU)}" cy="{int(tall * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            # Centred, which is the whole question.
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="ctr"/><a:lstStyle/>'
            f"{one}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "centre.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_DUMP_BLOCK="1", OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    told = []
    for line in done.stderr.decode("utf-8", "replace").splitlines():
        found = re.search(r"^block area (\d+)\.\.(\d+).*?block=([0-9.]+).*? at=([0-9.]+)", line)
        if found:
            told.append(tuple(float(one) for one in found.groups()))
    return ours, heights, lane, told


def foot(picture, top, bottom, lane):
    """The last row of ink — for `A`, one above the baseline."""
    held = (picture[top:bottom, lane:] < 128).sum(axis=1)
    lit = np.where(held > 0)[0]
    return top + int(lit[-1]) if len(lit) else None


def head(picture, top, bottom, lane):
    """The FIRST row of ink. A block of many lines has its foot governed by
    the leading as well as by the division, so the head is the cleaner reading
    of where the block was put."""
    held = (picture[top:bottom, lane:] < 128).sum(axis=1)
    lit = np.where(held > 0)[0]
    return top + int(lit[0]) if len(lit) else None


def main():
    global STAIRS
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    parser.add_argument("--staircase",
                        help="one face's name: sweep the box height at each "
                             "percentage instead of the usual three groups")
    parser.add_argument("--points", type=float, default=12.0)
    parser.add_argument("--pcts", default="70,90,115",
                        help="percentages to walk the height at; `a:b:step` "
                             "sweeps a range instead of naming them")
    parser.add_argument("--heights",
                        help="only these box heights, for sweeping the "
                             "percentage finely at a fixed height instead")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")
    if args.staircase:
        if ":" in args.pcts:
            first, last, step = (float(one) for one in args.pcts.split(":"))
            count = int(round((last - first) / step)) + 1
            pcts = [first + at * step for at in range(count)]
        else:
            pcts = [float(one) for one in args.pcts.split(",")]
        tall = ([int(one) for one in args.heights.split(",")]
                if args.heights else HEIGHTS)
        # Height first, so each height's run of percentages reads as one block:
        # at a fixed height the division's phase is fixed too, and where the
        # answer STEPS as the percentage grows is what says how far off `d` is.
        STAIRS = [(args.staircase, args.points, high, int(round(pct * 1000)), 1)
                  for high in tall for pct in pcts]

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane, told = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"{'face':<14}{'box':>5}{'pct':>7}{'n':>3} {'area':>12}{'block':>9}{'slack':>8}"
          f"{'Excel':>7}{'ours':>6}{'head E':>8}{'head O':>7}"
          f"{'round':>7}{'floor':>7}{'ceil':>6}")
    seen = {}
    for index, (face, points, tall, pct, lines) in enumerate(cases()):
        top = edges.get(index * SPACING)
        bottom = edges.get(index * SPACING + SPACING - 1)
        if top is None or bottom is None or index >= len(told):
            continue
        if bottom > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = foot(truth, top, bottom, lane)
        mine_foot = foot(mine, top, bottom, lane)
        if theirs is None or mine_foot is None:
            continue
        area_top, area_foot, block, _at = told[index]
        slack = (area_foot - area_top) - block
        # What each rule would put the block's top at, and so — since the
        # leading is the same in all three — how the ink foot would move.
        base = {name: area_top + rule(slack / 2.0)
                for name, rule in (("round", lambda v: np.floor(v + 0.5)),
                                   ("floor", np.floor), ("ceil", np.ceil))}
        shown = {name: theirs - (mine_foot - (_at - value))
                 for name, value in base.items()}
        got = [name for name, value in base.items() if value == _at]
        seen.setdefault(face, []).append(
            (tall, theirs - mine_foot, {name: int(value - _at) for name, value in base.items()}))
        their_head = head(truth, top, bottom, lane)
        my_head = head(mine, top, bottom, lane)
        print(f"{face:<14}{tall:>5}{(pct or 0) // 1000:>7}{lines:>3}"
              f" {f'{area_top:.0f}..{area_foot:.0f}':>12}{block:>9.3f}"
              f"{slack:>8.3f}{theirs:>7}{mine_foot:>6}"
              f"{their_head if their_head is not None else -1:>8}"
              f"{my_head if my_head is not None else -1:>7}"
              f"{int(base['round'] - _at):>7}{int(base['floor'] - _at):>7}"
              f"{int(base['ceil'] - _at):>6}"
              f"{'' if theirs == mine_foot else '  <<'}")
    print("\nExcel less ours, by face — a rule fits when its column below "
          "matches this everywhere")
    for face, rows in seen.items():
        print(f"  {face:<14}{[one[1] for one in rows]}")
        for name in ("round", "floor", "ceil"):
            print(f"    {name:<12}{[one[2][name] for one in rows]}")


if __name__ == "__main__":
    main()
