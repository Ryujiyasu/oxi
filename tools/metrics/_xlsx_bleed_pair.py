# -*- coding: utf-8 -*-
r"""When does a cell's ink cross the line into the row below?

Two books disagree. In `_xlsx_valign_pixels.py` the last scanline of a
top-aligned line lands one pixel into the row below; in `_xlsx_row_bleed.py`,
where the neighbours are tall and empty, the ink stops dead at the row's own
edge. This walks the difference one property at a time: the row's height, what
the row below holds, and whether the text has a descender.

Each case is a pair of rows with a tall empty spacer above it. The case text
sits in column A, the follower's in column C, so the case row's own ink is
told apart by where it sits across the sheet.

    python tools\metrics\_xlsx_bleed_pair.py
"""
import argparse
import os
import subprocess
import sys
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_bleed_pair")
BOOK = SCRATCH / "pair.xlsx"

HEIGHTS = [18.0, 13.5, 12.75, 12.0, 11.25, 10.5, 9.0]   # points: 24 … 12 px
FOLLOWERS = ["centre", "empty", "top"]
TEXTS = ["あA", "あAg"]
FACE, POINTS = "ＭＳ Ｐゴシック", 11.0          # line box 18 px, baseline 16
SPACER = 24.0


def build():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    SCRATCH.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 12.0
    sheet.column_dimensions["B"].width = 3.0
    sheet.column_dimensions["C"].width = 12.0
    cases, row = [], 1
    for height in HEIGHTS:
        for follower in FOLLOWERS:
            for text in TEXTS:
                sheet.row_dimensions[row].height = SPACER
                cell = sheet.cell(row=row + 1, column=1, value=text)
                cell.font = Font(name=FACE, size=POINTS)
                cell.alignment = Alignment(vertical="top", horizontal="left")
                sheet.row_dimensions[row + 1].height = height
                sheet.row_dimensions[row + 2].height = height
                if follower != "empty":
                    below = sheet.cell(row=row + 2, column=3, value=text)
                    below.font = Font(name=FACE, size=POINTS)
                    below.alignment = Alignment(
                        vertical="center" if follower == "centre" else "top",
                        horizontal="left")
                cases.append((row + 1, height, follower, text))
                row += 3
    book.save(BOOK)
    return cases


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=600)
    listing.unlink(missing_ok=True)
    return picture


def draw():
    ours = SCRATCH / "pair.oxi.png"
    environment = dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1")
    done = subprocess.run([str(RENDERER), str(BOOK), str(ours), "96"],
                          capture_output=True, timeout=300, env=environment)
    heights, columns = {}, {}
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column":
            columns[int(parts[1])] = int(float(parts[3]))
    return ours, heights, columns


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()

    cases = build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    ours_png, heights, columns = draw()
    truth = np.asarray(Image.open(picture).convert("L"))
    ours = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = (at, at + heights[index])
        at += heights[index]
    lane = columns.get(min(columns), 0) if columns else 0

    print(f"{'row px':>7}{'follower':>10}{'text':>6}"
          f"{'Excel foot':>12}{'ours foot':>11}   (pixels past the row's last)")
    for case_row, height, follower, text in cases:
        if case_row + 1 not in edges:
            continue
        top, foot = edges[case_row]
        stop = min(edges[case_row + 1][1], truth.shape[0], ours.shape[0])

        def past(image):
            band = (image[top:stop, :lane] < 128).sum(axis=1)
            lit = np.flatnonzero(band)
            return top + int(lit[-1]) - (foot - 1) if lit.size else None

        print(f"{foot - top:>7}{follower:>10}{text:>6}"
              f"{str(past(truth)):>12}{str(past(ours)):>11}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
