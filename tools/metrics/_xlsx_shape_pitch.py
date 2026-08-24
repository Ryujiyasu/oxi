# -*- coding: utf-8 -*-
r"""How far apart does Excel set the lines of a shape's text?

`dc4fcff7f5f8_001` draws the same words as Excel on every line and drifts down
the block, which is a pitch that is close but not equal. Its panels are メイリオ
at four sizes with `<a:lnSpc>` pinned in points on some paragraphs and stated
as a percentage on others.

Each case is a shape holding five lines of one face, size and spacing. What is
read back is where each line's ink starts, so both the first line's place and
the step between them can be held against Excel's.

    python tools\metrics\_xlsx_shape_pitch.py
    python tools\metrics\_xlsx_shape_pitch.py --reuse
"""
import argparse
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import numpy as np
from PIL import Image

REPO = Path(__file__).resolve().parents[2]
SHOOTER = Path(__file__).resolve().parent / "_xlsx_screen_shot.ps1"
RENDERER = REPO / "tools" / "oxi-xlsx-renderer" / "target" / "release" / "oxi-xlsx-renderer.exe"
SCRATCH = Path(r"C:\tmp\xlsx_shape_pitch")
BOOK = SCRATCH / "pitch.xlsx"
EMU = 9525.0

SPACING = 16            # rows per case, tall enough that nothing overflows
LINES = 4
# A letter that sits on the baseline: its ink foot is the baseline, so the
# question "where does Excel put the line" is answered without having to know
# the face's ascent first.
WORD = "A"

# メイリオ is the one face whose first baseline does not sit at 0.750 of a
# pinned pitch, so it is swept at four sizes against three faces that do (and
# against 游ゴシック, which the earlier note called a mild deviant). The
# question is what metric of the face predicts the residual.
FONTS = [("Yu Gothic UI", 12.0), ("Yu Gothic UI", 11.0), ("メイリオ", 11.0),
         ("ＭＳ Ｐゴシック", 11.0), ("游ゴシック", 11.0)]
# The two anchors gave the same first baseline in every one of the earlier
# thirty-two rows, so only the top one is worth the rows.
ANCHORS = ["t"]
# What `<a:lnSpc>` can say, and saying nothing at all.
SPACINGS = [("none", "")]
# A fine sweep of the pinned pitch: the slope of the first baseline against
# the pitch is what says where in the line Excel puts it.
SPACINGS += [(f"pts {value}", f'<a:lnSpc><a:spcPts val="{value}"/></a:lnSpc>')
             for value in (1500, 2100, 3000)]
# `glossary_05` sets its flowchart in Yu Gothic UI 12pt at 80% — a PERCENTAGE,
# which the pinned sweep never covered.
SPACINGS += [(f"pct {value}", f'<a:lnSpc><a:spcPct val="{value}"/></a:lnSpc>')
             for value in (70000, 80000, 90000, 100000, 115000, 150000)]


def cases():
    return [(face, points, f"{name} {anchor}", (spec, anchor))
            for face, points in FONTS
            for name, spec in SPACINGS
            for anchor in ANCHORS]


def anchors_xml():
    held = []
    for index, (face, points, _name, told) in enumerate(cases()):
        spec, anchor = told
        runs = "".join(
            f'<a:p><a:pPr algn="l">{spec}</a:pPr><a:r>'
            f'<a:rPr lang="ja-JP" sz="{int(points * 100)}">'
            f'<a:latin typeface="{face}"/><a:ea typeface="{face}"/>'
            f"</a:rPr><a:t>{WORD}</a:t></a:r></a:p>"
            for _ in range(LINES))
        held.append(
            f"<xdr:oneCellAnchor>"
            f"<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>{index * SPACING}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f'<xdr:ext cx="{int(200 * EMU)}" cy="{int(15 * 20 * EMU)}"/>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="box {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr>"
            f'<xdr:spPr><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f"<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>"
            f'<xdr:txBody><a:bodyPr wrap="square" anchor="t"/><a:lstStyle/>'
            f"{runs}</xdr:txBody></xdr:sp><xdr:clientData/></xdr:oneCellAnchor>"
        )
    return "".join(held)


def build():
    from openpyxl import Workbook

    SCRATCH.mkdir(parents=True, exist_ok=True)
    plain = SCRATCH / "_plain.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.column_dimensions["A"].width = 2.0
    sheet.column_dimensions["B"].width = 40.0
    sheet.cell(row=1, column=1, value="a")
    sheet.cell(row=len(cases()) * SPACING + 2, column=4, value="z")
    book.save(plain)

    drawing = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        f"{anchors_xml()}</xdr:wsDr>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1"'
        ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"'
        ' Target="../drawings/drawing1.xml"/></Relationships>'
    )
    BOOK.unlink(missing_ok=True)
    with zipfile.ZipFile(plain) as source, \
            zipfile.ZipFile(BOOK, "w", zipfile.ZIP_DEFLATED) as out:
        for item in source.infolist():
            held = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                held = held.decode("utf-8").replace(
                    "</Types>",
                    '<Override PartName="/xl/drawings/drawing1.xml"'
                    ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
                    "</Types>",
                ).encode("utf-8")
            if item.filename == "xl/worksheets/sheet1.xml":
                held = held.decode("utf-8").replace(
                    "</worksheet>", '<drawing r:id="rId1"/></worksheet>')
                if "xmlns:r=" not in held:
                    held = held.replace(
                        "<worksheet ",
                        '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships" ', 1)
                held = held.encode("utf-8")
            out.writestr(item, held)
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)


def shoot():
    picture = BOOK.with_suffix(".excel.png")
    picture.unlink(missing_ok=True)
    listing = SCRATCH / "_batch.txt"
    listing.write_text(f"{BOOK.resolve()}\t{picture.resolve()}", encoding="utf-8-sig")
    subprocess.run(["powershell", "-NoProfile", "-File", str(SHOOTER),
                    "-ListFile", str(listing.resolve())],
                   capture_output=True, text=True, encoding="utf-8",
                   errors="replace", timeout=1800)
    listing.unlink(missing_ok=True)
    return picture


def drawn():
    ours = SCRATCH / "pitch.oxi.png"
    done = subprocess.run(
        [str(RENDERER), str(BOOK), str(ours), "96"], capture_output=True, timeout=1800,
        env=dict(os.environ, OXI_XLSX_DUMP_ROWS="1", OXI_XLSX_DUMP_COLUMNS="1",
                 OXI_XLSX_SHAPE_TEXT="1"))
    heights, lane = {}, 0
    for line in done.stdout.decode("utf-8", "replace").splitlines():
        parts = line.split()
        if len(parts) == 4 and parts[0] == "row":
            heights[int(parts[1])] = int(float(parts[3]))
        if len(parts) == 4 and parts[0] == "column" and lane == 0:
            lane = int(float(parts[3]))
    return ours, heights, lane


def feet(picture, top, foot, lane):
    held = (picture[top:foot, lane:] < 128).sum(axis=1)
    lines, run = [], None
    for step, lit in enumerate(held):
        if lit:
            run = [step, step] if run is None else [run[0], step]
        elif run is not None and step - run[1] > 2:
            lines.append(run[1])
            run = None
    if run is not None:
        lines.append(run[1])
    return lines


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--reuse", action="store_true")
    args = parser.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    build()
    picture = BOOK.with_suffix(".excel.png") if args.reuse else shoot()
    if not picture.exists():
        print("Excel gave no picture")
        return
    truth = np.asarray(Image.open(picture).convert("L"))
    ours_png, heights, lane = drawn()
    mine = np.asarray(Image.open(ours_png).convert("L"))
    edges, at = {}, 0
    for index in sorted(heights):
        edges[index] = at
        at += heights[index]

    print(f"{'face':<14}{'pt':>5}{'lnSpc':>14}"
          f"{'Excel base1':>12}{'pitch':>8}{'ours base1':>12}{'pitch':>8}{'lines':>7}")
    for index, (face, points, name, _held) in enumerate(cases()):
        top = edges.get(index * SPACING)
        foot = edges.get(index * SPACING + SPACING - 1)
        if top is None or foot is None or foot > min(truth.shape[0], mine.shape[0]):
            continue
        theirs = feet(truth, top, foot, lane)
        ours = feet(mine, top, foot, lane)
        step = lambda held: (f"{(held[-1] - held[0]) / (len(held) - 1):.2f}"
                             if len(held) > 1 else "-")
        flag = "" if (theirs[:1] == ours[:1] and step(theirs) == step(ours)) else "  <<"
        print(f"{face:<14}{points:>5}{name:>14}"
              f"{theirs[0] if theirs else -1:>12}{step(theirs):>8}"
              f"{ours[0] if ours else -1:>12}{step(ours):>8}"
              f"{f'{len(theirs)}/{len(ours)}':>7}{flag}")


if __name__ == "__main__":
    main()
